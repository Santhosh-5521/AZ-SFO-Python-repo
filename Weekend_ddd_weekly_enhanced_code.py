import pandas as pd
import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
import os 
from os import listdir
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import shutil
from shutil import make_archive

prid = input("enter your prid:")

input_xpo_w_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\input\f_sls_hcp_prod_plan_wk.xlsx" %(prid))

xpo_w_nrx = input_xpo_w_df.pivot_table(index = ["mkt_nm","brd_nm","prod_nm"],values = ["nrx_new","nrx_old"])
xpo_w_nrx["deviation"] = ((xpo_w_nrx.nrx_new - xpo_w_nrx.nrx_old)/xpo_w_nrx.nrx_old)*100
#xpo_w_nrx
xpo_w_nrx["deviation"] = xpo_w_nrx["deviation"].round(round(1))
#xpo_w_nrx
xpo_w_nrx_deviation = xpo_w_nrx[((xpo_w_nrx['deviation'] <= -5) | (xpo_w_nrx['deviation'] >=5))]
#xpo_w_nrx_deviation
xpo_w_nrx_deviations = xpo_w_nrx_deviation[((xpo_w_nrx_deviation['nrx_new'] != 0) & (xpo_w_nrx_deviation['nrx_old'] != 0))]
#xpo_w_nrx_deviations
xpo_w_nrx_zero_sales = xpo_w_nrx[(xpo_w_nrx['nrx_new'] == 0) | (xpo_w_nrx['nrx_old'] == 0)]
#xpo_w_nrx_zero_sales

xpo_w_nrx.to_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\xpo_w_nrx.xlsx" %(prid),sheet_name = "xpo_w_nrx")

xpo_w_nrx_deviations_flag = 'N'
xpo_w_nrx_zero_sales_flag = 'N'
xpo_w_trx_deviations_flag = 'N'
xpo_w_trx_zero_sales_flag = 'N'

if xpo_w_nrx_deviations.empty == True:
 xpo_w_nrx_deviations_flag = 'Y'
else:
 xpo_w_nrx_deviations.to_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\deviations_nrx.xlsx" %(prid),sheet_name = "deviations_nrx")

if xpo_w_nrx_zero_sales.empty == True:
 xpo_w_nrx_zero_sales_flag = 'Y'
else:
 xpo_w_nrx_zero_sales.to_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\zero_sales_nrx.xlsx" %(prid),sheet_name = "zero_sales_nrx")

##Zero_sales_nrx:

if xpo_w_nrx_zero_sales_flag == 'Y':
  pass
else:
 zs_xpo_w_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\zero_sales_nrx.xlsx" %(prid))
 xpo_w_zero_sales_nrx = zs_xpo_w_df.fillna(method = "ffill")
#xpo_w_zero_sales_nrx

 xpo_w_zero_sales_nrx.loc[(xpo_w_zero_sales_nrx['nrx_old'] == 0) & (xpo_w_zero_sales_nrx['nrx_new'] == 0), 'comments'] = "no comments"
 xpo_w_zero_sales_nrx.loc[(xpo_w_zero_sales_nrx['nrx_new'] == 0) & (xpo_w_zero_sales_nrx['nrx_old'] != 0), 'comments'] = xpo_w_zero_sales_nrx['prod_nm'] + " sales dropped to zero under " + xpo_w_zero_sales_nrx["mkt_nm"]
 xpo_w_zero_sales_nrx.loc[(xpo_w_zero_sales_nrx['nrx_old'] == 0) & (xpo_w_zero_sales_nrx['nrx_new'] != 0), 'comments'] = xpo_w_zero_sales_nrx['prod_nm'] + " is flowing under " + xpo_w_zero_sales_nrx["mkt_nm"]
 
 xpo_w_zero_sales_nrx = xpo_w_zero_sales_nrx.sort_values(by='prod_nm')

##Deviation_nrx:

if xpo_w_nrx_deviations_flag == 'Y':
  pass
else:
 dev_xpo_w_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\deviations_nrx.xlsx" %(prid))
 xpo_w_deviation_nrx = dev_xpo_w_df.fillna(method = "ffill")
#xpo_w_deviation_nrx

 xpo_w_deviation_nrx.loc[xpo_w_deviation_nrx['deviation'] > 5, 'comments'] = xpo_w_deviation_nrx['brd_nm'] + " sales is increased by " + xpo_w_deviation_nrx['deviation'].map(str) + "%" + " in " + xpo_w_deviation_nrx['mkt_nm']
 xpo_w_deviation_nrx.loc[xpo_w_deviation_nrx['deviation'] < -5, 'comments'] = xpo_w_deviation_nrx['brd_nm'] + " sales is dropped by " + xpo_w_deviation_nrx['deviation'].map(str) + "%" + " in " + xpo_w_deviation_nrx['mkt_nm']
 
 xpo_w_deviation_nrx = xpo_w_deviation_nrx.sort_values(by='prod_nm')

#############################################Trx##############################################################################################

xpo_w_trx = input_xpo_w_df.pivot_table(index = ["mkt_nm","brd_nm","prod_nm"],values = ["trx_new","trx_old"])
xpo_w_trx["deviation"] = ((xpo_w_trx.trx_new - xpo_w_trx.trx_old)/xpo_w_trx.trx_old)*100
#xpo_w_trx
xpo_w_trx["deviation"] = xpo_w_trx["deviation"].round(round(1))
#xpo_w_trx
xpo_w_trx_deviation = xpo_w_trx[((xpo_w_trx['deviation'] <= -5) | (xpo_w_trx['deviation'] >=5))]
#xpo_w_trx_deviation
xpo_w_trx_deviations = xpo_w_trx_deviation[((xpo_w_trx_deviation['trx_new'] != 0) & (xpo_w_trx_deviation['trx_old'] != 0))]
#xpo_w_trx_deviations
xpo_w_trx_zero_sales = xpo_w_trx[(xpo_w_trx['trx_new'] == 0) | (xpo_w_trx['trx_old'] == 0)]
#xpo_w_trx_zero_sales

xpo_w_trx.to_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\xpo_w_trx.xlsx" %(prid),sheet_name = "xpo_w_trx")

if xpo_w_trx_deviations.empty == True:
 xpo_w_trx_deviations_flag = 'Y'
else:
 xpo_w_trx_deviations.to_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\deviations_trx.xlsx" %(prid),sheet_name = "deviations_trx")
 
if xpo_w_trx_zero_sales.empty == True:
 xpo_w_trx_zero_sales_flag = 'Y'
else:
 xpo_w_trx_zero_sales.to_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\zero_sales_trx.xlsx" %(prid),sheet_name = "zero_sales_trx")

##Zero_sales_trx:

if xpo_w_trx_zero_sales_flag == 'Y':
  pass
else:
 Zero_sales_trx_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\zero_sales_trx.xlsx" %(prid))
 xpo_w_zero_sales_trx = Zero_sales_trx_df.fillna(method = "ffill")
#zero_sales_trx

 xpo_w_zero_sales_trx.loc[(xpo_w_zero_sales_trx['trx_old'] == 0) & (xpo_w_zero_sales_trx['trx_new'] == 0), 'comments'] = "no comments"
 xpo_w_zero_sales_trx.loc[(xpo_w_zero_sales_trx['trx_new'] == 0) & (xpo_w_zero_sales_trx['trx_old'] != 0), 'comments'] = xpo_w_zero_sales_trx['prod_nm'] + " sales dropped to zero under " + xpo_w_zero_sales_trx["mkt_nm"]
 xpo_w_zero_sales_trx.loc[(xpo_w_zero_sales_trx['trx_old'] == 0) & (xpo_w_zero_sales_trx['trx_new'] != 0), 'comments'] = xpo_w_zero_sales_trx['prod_nm'] + " is flowing under " + xpo_w_zero_sales_trx["mkt_nm"]
 
 xpo_w_zero_sales_trx = xpo_w_zero_sales_trx.sort_values(by='prod_nm')

##Deviation_trx:

if xpo_w_trx_deviations_flag == 'Y':
  pass
else:
 Deviation_trx_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\deviations_trx.xlsx" %(prid))
 xpo_w_deviation_trx = Deviation_trx_df.fillna(method = "ffill")
#xpo_w_deviation_trx

#deviation_trx['deviation']= deviation_trx['deviation'].astype(str)

 xpo_w_deviation_trx.loc[xpo_w_deviation_trx['deviation'] > 5, 'comments'] = xpo_w_deviation_trx['brd_nm'] + " sales is increased by " + xpo_w_deviation_trx['deviation'].map(str) + "%" + " in " + xpo_w_deviation_trx['mkt_nm']
 xpo_w_deviation_trx.loc[xpo_w_deviation_trx['deviation'] < -5, 'comments'] = xpo_w_deviation_trx['brd_nm'] + " sales is dropped by " + xpo_w_deviation_trx['deviation'].map(str) + "%" + " in " + xpo_w_deviation_trx['mkt_nm']
 xpo_w_deviation_trx = xpo_w_deviation_trx.sort_values(by='prod_nm')
 
####################################Wrting the output to excel file################################################################### 	

if [xpo_w_nrx_deviations_flag == 'Y'] or [xpo_w_nrx_zero_sales_flag == 'Y'] or [xpo_w_trx_deviations_flag == 'Y'] or [xpo_w_trx_zero_sales_flag == 'Y']:
    with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx" %(prid)) as writer:    
      xpo_w_nrx.to_excel(writer,sheet_name = "xpo_w_nrx")
      xpo_w_trx.to_excel(writer,sheet_name = "xpo_w_trx")
    
    if xpo_w_nrx_deviations_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx" %(prid))
        book.create_sheet("deviation_nrx")
        sheet = book["deviation_nrx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx" %(prid))
        
    else:
        with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx" %(prid), engine='openpyxl', mode='a') as writer: 
         xpo_w_deviation_nrx.to_excel(writer,sheet_name = "deviation_nrx")

    if xpo_w_nrx_zero_sales_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx" %(prid))
        book.create_sheet("Zero_sales_nrx")
        sheet = book["Zero_sales_nrx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx" %(prid))
        
    else:
        with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx" %(prid), engine='openpyxl', mode='a') as writer: 
         xpo_w_zero_sales_nrx.to_excel(writer,sheet_name = "zero_sales_nrx")
   
    if xpo_w_trx_deviations_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx" %(prid))
        book.create_sheet("deviation_trx")
        sheet = book["deviation_trx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx" %(prid))
        
    else:
        with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx" %(prid), engine='openpyxl', mode='a') as writer: 
         xpo_w_deviation_trx.to_excel(writer,sheet_name = "deviation_trx")
  

    if xpo_w_trx_zero_sales_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx" %(prid))
        book.create_sheet("Zero_sales_trx")
        sheet = book["Zero_sales_trx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx" %(prid))
        
    else:
        with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx" %(prid), engine='openpyxl', mode='a') as writer: 
         xpo_w_zero_sales_trx.to_excel(writer,sheet_name = "Zero_sales_trx")
		 
##################################################Data Visualization#########################################################################

####################################################MARKET##################################################################################

mkt_dev = xpo_w_deviation_nrx[['mkt_nm', 'deviation']].copy()
#mkt_dev

mkt_dev.to_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\mkt_deviation.xlsx" %(prid),index=False)
	  
market_deviation = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\mkt_deviation.xlsx" %(prid))
market_deviation

mar_res = market_deviation.groupby('mkt_nm')
#mar_res

mar_out = mar_res.mean()
#mar_out

mar_out.to_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\mkt_dev_output.xlsx" %(prid))

mar_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\mkt_dev_output.xlsx" %(prid))
	  
sns.set(rc={'figure.figsize':(11.7,8.27)})
mar_sns_plot = sns.barplot(x="deviation",y="mkt_nm",data=mar_df)
mar_sns_plot.figure.savefig(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\mar_output.pdf" %(prid),bbox_inches="tight",pad_inches=1)

#######################################BRAND######################################################################################

brand_dev = xpo_w_deviation_nrx[['brd_nm', 'deviation']].copy()
#brand_dev

brand_dev.to_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\brand_deviation.xlsx" %(prid),index=False)

brand_deviation = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\brand_deviation.xlsx" %(prid))
brand_deviation

brand_res = brand_deviation.groupby('brd_nm')
#brand_res

brand_out = brand_res.mean()
#brand_out

brand_out.to_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\brand_dev_output.xlsx" %(prid))

brand_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\brand_dev_output.xlsx" %(prid))

sns.set(rc={'figure.figsize':(11.7,8.27)})
brand_sns_plot = sns.barplot(x="deviation",y="brd_nm",data=brand_df)
brand_sns_plot.figure.savefig(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\brand_output.pdf" %(prid),bbox_inches="tight",pad_inches=1)

###############################################PRODUCT##########################################################################

product_dev = xpo_w_deviation_nrx[['prod_nm', 'deviation']].copy()
#product_dev

product_dev.to_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\product_deviation.xlsx" %(prid),index=False)

product_deviation = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\product_deviation.xlsx" %(prid))
product_deviation

product_res = product_deviation.groupby('prod_nm')
#product_res

product_out = product_res.mean()
#product_out

product_out.to_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\product_dev_output.xlsx" %(prid))

product_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\product_dev_output.xlsx" %(prid))

sns.set(rc={'figure.figsize':(11.7,8.27)})
prod_sns_plot = sns.barplot(x="deviation",y="prod_nm",data=product_df)
prod_sns_plot.figure.savefig(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\product_output.pdf" %(prid),bbox_inches="tight",pad_inches=1)

#################################Moving the image files to output folder######################################################

sourcepath = r'C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\\' %(prid)
sourcefiles = os.listdir(sourcepath)
destinationpath = r'C:\Users\%s\Desktop\unaligned\xpo_weekly\output\\' %(prid)
for file in sourcefiles:
    if file.endswith('.pdf'):
        shutil.move(os.path.join(sourcepath,file), os.path.join(destinationpath,file))
		

#####################################Zipping the output files##########################################################

dir_name = r'C:\Users\%s\Desktop\unaligned\xpo_weekly\output\\' %(prid)
destination = r'C:\Users\%s\Desktop\unaligned\xpo_weekly\zip\\' %(prid)
shutil.make_archive("xpo_w_output", 'zip', dir_name)
shutil.move('%s.%s'%("xpo_w_output",'zip'), destination)

sourcepath = r'C:\Users\%s\Desktop\unaligned\xpo_weekly\output\\' %(prid)
sourcefiles = os.listdir(sourcepath)
destinationpath = r'C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\\' %(prid)
for file in sourcefiles:
    if file.endswith('.pdf'):
        shutil.move(os.path.join(sourcepath,file), os.path.join(destinationpath,file))

###################################################Deleting unwanted files after run from temp folder########################################################

import os
import shutil

for root, dirs, files in os.walk(r'C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\\' %(prid)):
    for f in files:
        os.unlink(os.path.join(root, f))
    for d in dirs:
        shutil.rmtree(os.path.join(root, d))


