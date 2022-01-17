#!/usr/bin/env python
# coding: utf-8

# In[1]:


#!/usr/bin/env python
# coding: utf-8

# In[1]:


#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
import os 
from os import listdir
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import shutil
from shutil import make_archive

prid = input("enter your prid:")

######################Reading input files from local and storing it in dataframes##############################################################

input_ddd_w_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\input\f_sls_outl_prod_wk.xlsx" %(prid))
input_ddd_m_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\input\f_sls_outl_prod_mnth.xlsx" %(prid))
input_xpo_w_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\input\f_sls_hcp_prod_plan_wk.xlsx" %(prid))
input_xpopd_w_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\input\f_sls_hcp_prod_plan_dyn_wk.xlsx" %(prid))
input_xpopd_m_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\input\f_sls_hcp_prod_plan_dyn_mnth.xlsx" %(prid))

###################################################DDD WEEKLY############################################################################################

##########################UNITS#############################################################################

ddd_w_units = input_ddd_w_df.pivot_table(index = ["mkt_nm","brd_nm","prod_nm"],values = ["tot_units_new","tot_units_old"])
ddd_w_units["deviation"] = ((ddd_w_units.tot_units_new - ddd_w_units.tot_units_old)/ddd_w_units.tot_units_old)*100
#ddd_w_units
ddd_w_units["deviation"] = ddd_w_units["deviation"].round(round(1))
#ddd_w_units
ddd_w_units_deviation = ddd_w_units[((ddd_w_units['deviation'] <= -5) | (ddd_w_units['deviation'] >=5))]
#ddd_w_units_deviation
ddd_w_units_deviations = ddd_w_units_deviation[((ddd_w_units_deviation['tot_units_new'] != 0) & (ddd_w_units_deviation['tot_units_old'] != 0))]
#ddd_w_units_deviations
ddd_w_units_zero_sales = ddd_w_units[(ddd_w_units['tot_units_new'] == 0) | (ddd_w_units['tot_units_old'] == 0)]
#ddd_w_units_zero_sales

ddd_w_units.to_excel(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\ddd_w_units.xlsx" %(prid),sheet_name = "ddd_w_units")

ddd_w_units_deviations_flag = 'N'
ddd_w_units_zero_sales_flag = 'N'
ddd_w_dollars_deviations_flag = 'N'
ddd_w_dollars_zero_sales_flag = 'N'

if ddd_w_units_deviations.empty == True:
 ddd_w_units_deviations_flag = 'Y'
else:
 ddd_w_units_deviations.to_excel(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\deviations_units.xlsx" %(prid),sheet_name = "deviations_units")

if ddd_w_units_zero_sales.empty == True:
 ddd_w_units_zero_sales_flag = 'Y'
else:
 ddd_w_units_zero_sales.to_excel(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\zero_sales_units.xlsx" %(prid),sheet_name = "zero_sales_units")

##Zero_sales_units:

if ddd_w_units_deviations_flag == 'Y':
  pass
else:
 zs_ddd_w_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\zero_sales_units.xlsx" %(prid))
 ddd_w_zero_sales_units = zs_ddd_w_df.fillna(method = "ffill")
#zero_sales_units

 ddd_w_zero_sales_units.loc[(ddd_w_zero_sales_units['tot_units_old'] == 0) & (ddd_w_zero_sales_units['tot_units_new'] == 0), 'comments'] = "no comments"
 ddd_w_zero_sales_units.loc[(ddd_w_zero_sales_units['tot_units_new'] == 0) & (ddd_w_zero_sales_units['tot_units_old'] != 0), 'comments'] = ddd_w_zero_sales_units['prod_nm'] + " sales dropped to zero under " + ddd_w_zero_sales_units["mkt_nm"]
 ddd_w_zero_sales_units.loc[(ddd_w_zero_sales_units['tot_units_old'] == 0) & (ddd_w_zero_sales_units['tot_units_new'] != 0), 'comments'] = ddd_w_zero_sales_units['prod_nm'] + " is flowing under " + ddd_w_zero_sales_units["mkt_nm"]

 ddd_w_zero_sales_units = ddd_w_zero_sales_units.sort_values(by='prod_nm')
 
##Deviation_units:

if ddd_w_units_zero_sales_flag == 'Y':
  pass
else:
 dev_ddd_w_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\deviations_units.xlsx" %(prid))
 ddd_w_deviation_units = dev_ddd_w_df.fillna(method = "ffill")
#ddd_w_deviation_units

 ddd_w_deviation_units.loc[ddd_w_deviation_units['deviation'] > 5, 'comments'] = ddd_w_deviation_units['brd_nm'] + " sales is increased by " + ddd_w_deviation_units['deviation'].map(str) + "%" + " in " + ddd_w_deviation_units['mkt_nm']
 ddd_w_deviation_units.loc[ddd_w_deviation_units['deviation'] < -5, 'comments'] = ddd_w_deviation_units['brd_nm'] + " sales is dropped by " + ddd_w_deviation_units['deviation'].map(str) + "%" + " in " + ddd_w_deviation_units['mkt_nm']
 
 ddd_w_deviation_units = ddd_w_deviation_units.sort_values(by='prod_nm')

#############################################DOLLARS##############################################################################################

ddd_w_dollars = input_ddd_w_df.pivot_table(index = ["mkt_nm","brd_nm","prod_nm"],values = ["tot_dolrs_new","tot_dolrs_old"])
ddd_w_dollars["deviation"] = ((ddd_w_dollars.tot_dolrs_new - ddd_w_dollars.tot_dolrs_old)/ddd_w_dollars.tot_dolrs_old)*100
#ddd_w_dollars
ddd_w_dollars["deviation"] = ddd_w_dollars["deviation"].round(round(1))
#ddd_w_dollars
ddd_w_dollars_deviation = ddd_w_dollars[((ddd_w_dollars['deviation'] <= -5) | (ddd_w_dollars['deviation'] >=5))]
#ddd_w_dollars_deviation
ddd_w_dollars_deviations = ddd_w_dollars_deviation[((ddd_w_dollars_deviation['tot_dolrs_new'] != 0) & (ddd_w_dollars_deviation['tot_dolrs_old'] != 0))]
#ddd_w_dollars_deviations
ddd_w_dollars_zero_sales = ddd_w_dollars[(ddd_w_dollars['tot_dolrs_new'] == 0) | (ddd_w_dollars['tot_dolrs_old'] == 0)]
#ddd_w_dollars_zero_sales

ddd_w_dollars.to_excel(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\ddd_w_dollars.xlsx" %(prid),sheet_name = "ddd_w_dollars")

if ddd_w_dollars_deviations.empty == True:
 ddd_w_dollars_deviations_flag = 'Y'
else:
 ddd_w_dollars_deviations.to_excel(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\deviations_dollars.xlsx" %(prid),sheet_name = "deviations_dollars")
 
if ddd_w_dollars_zero_sales.empty == True:
 ddd_w_dollars_zero_sales_flag = 'Y'
else:
 ddd_w_dollars_zero_sales.to_excel(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\zero_sales_dollars.xlsx" %(prid),sheet_name = "zero_sales_dollars")

##Zero_sales_dollars:

if ddd_w_dollars_deviations_flag == 'Y':
  pass
else:
 Zero_sales_dollars_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\zero_sales_dollars.xlsx" %(prid))
 ddd_w_zero_sales_dollars = Zero_sales_dollars_df.fillna(method = "ffill")
#ddd_w_zero_sales_dollars

 ddd_w_zero_sales_dollars.loc[(ddd_w_zero_sales_dollars['tot_dolrs_old'] == 0) & (ddd_w_zero_sales_dollars['tot_dolrs_new'] == 0), 'comments'] = "no comments"
 ddd_w_zero_sales_dollars.loc[(ddd_w_zero_sales_dollars['tot_dolrs_new'] == 0) & (ddd_w_zero_sales_dollars['tot_dolrs_old'] != 0), 'comments'] = ddd_w_zero_sales_dollars['prod_nm'] + " sales dropped to zero under " + ddd_w_zero_sales_dollars["mkt_nm"]
 ddd_w_zero_sales_dollars.loc[(ddd_w_zero_sales_dollars['tot_dolrs_old'] == 0) & (ddd_w_zero_sales_dollars['tot_dolrs_new'] != 0), 'comments'] = ddd_w_zero_sales_dollars['prod_nm'] + " is flowing under " + ddd_w_zero_sales_dollars["mkt_nm"]
 
 ddd_w_zero_sales_dollars = ddd_w_zero_sales_dollars.sort_values(by='prod_nm')

##Deviation_dollars:

if ddd_w_dollars_zero_sales_flag == 'Y':
  pass
else:
 Deviation_dollars_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\deviations_dollars.xlsx" %(prid))
 ddd_w_deviation_dollars = Deviation_dollars_df.fillna(method = "ffill")
#deviation_dollars

#deviation_dollars['deviation']= deviation_dollars['deviation'].astype(str)
 
 ddd_w_deviation_dollars.loc[ddd_w_deviation_dollars['deviation'] > 5, 'comments'] = ddd_w_deviation_dollars['brd_nm'] + " sales is increased by " + ddd_w_deviation_dollars['deviation'].map(str) + "%" + " in " + ddd_w_deviation_dollars['mkt_nm']
 ddd_w_deviation_dollars.loc[ddd_w_deviation_dollars['deviation'] < -5, 'comments'] = ddd_w_deviation_dollars['brd_nm'] + " sales is dropped by " + ddd_w_deviation_dollars['deviation'].map(str) + "%" + " in " + ddd_w_deviation_dollars['mkt_nm']
 ddd_w_deviation_dollars = ddd_w_deviation_dollars.sort_values(by='prod_nm')
 
#########################Writing the output to excel file#########################################################################

if [ddd_w_units_deviations_flag == 'Y'] or [ddd_w_units_zero_sales_flag == 'Y'] or [ddd_w_dollars_deviations_flag == 'Y'] or [ddd_w_dollars_zero_sales_flag == 'Y']:
    with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\output\ddd_w_output.xlsx" %(prid)) as writer:    
      ddd_w_units.to_excel(writer,sheet_name = "ddd_w_units")
      ddd_w_dollars.to_excel(writer,sheet_name = "ddd_w_dollars")
    
    if ddd_w_units_deviations_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\output\ddd_w_output.xlsx" %(prid))
        book.create_sheet("deviation_units")
        sheet = book["deviation_units"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\output\ddd_w_output.xlsx" %(prid))
        
    else:
        with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\output\ddd_w_output.xlsx" %(prid), engine='openpyxl', mode='a') as writer: 
         ddd_w_deviation_units.to_excel(writer,sheet_name = "deviation_units")

    if ddd_w_units_zero_sales_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\output\ddd_w_output.xlsx" %(prid))
        book.create_sheet("Zero_sales_units")
        sheet = book["Zero_sales_units"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\output\ddd_w_output.xlsx" %(prid))
        
    else:
        with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\output\ddd_w_output.xlsx" %(prid), engine='openpyxl', mode='a') as writer: 
         ddd_w_zero_sales_units.to_excel(writer,sheet_name = "Zero_sales_units")  
    
    if ddd_w_dollars_deviations_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\output\ddd_w_output.xlsx" %(prid))
        book.create_sheet("deviation_dollars")
        sheet = book["deviation_dollars"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\output\ddd_w_output.xlsx" %(prid))
        
    else:
        with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\output\ddd_w_output.xlsx" %(prid), engine='openpyxl', mode='a') as writer: 
         ddd_w_deviation_dollars.to_excel(writer,sheet_name = "deviation_dollars")
  

    if ddd_w_dollars_zero_sales_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\output\ddd_w_output.xlsx" %(prid))
        book.create_sheet("Zero_sales_dollars")
        sheet = book["Zero_sales_dollars"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\output\ddd_w_output.xlsx" %(prid))
        
    else:
        with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\output\ddd_w_output.xlsx" %(prid), engine='openpyxl', mode='a') as writer: 
         ddd_w_zero_sales_dollars.to_excel(writer,sheet_name = "Zero_sales_dollars")
 
##################################################DDD_Weekly_Data_Visualization#########################################################################

####################################################MARKET##################################################################################

if ddd_w_units_deviations_flag == 'Y':
  pass
else:

 ddd_w_mkt_dev = ddd_w_deviation_units[['mkt_nm', 'deviation']].copy()
 #ddd_w_mkt_dev

 ddd_w_mkt_dev.to_excel(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\ddd_w_mkt_deviation.xlsx" %(prid),index=False)
 
 ddd_w_market_deviation = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\ddd_w_mkt_deviation.xlsx" %(prid))
 ddd_w_market_deviation

 ddd_w_mar_res = ddd_w_market_deviation.groupby('mkt_nm')
 #ddd_w_mar_res

 ddd_w_mar_out = ddd_w_mar_res.mean()
 #ddd_w_mar_out

 ddd_w_mar_out.to_excel(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\ddd_w_mkt_dev_output.xlsx" %(prid))

 ddd_w_mar_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\ddd_w_mkt_dev_output.xlsx" %(prid))
 
 sns.set(rc={'figure.figsize':(11.7,8.27)})
 ddd_w_mar_sns_plot = sns.barplot(x="deviation",y="mkt_nm",data=ddd_w_mar_df)
 ddd_w_mar_sns_plot.figure.savefig(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\ddd_w_mar_output.pdf" %(prid),bbox_inches="tight",pad_inches=1)

#######################################BRAND######################################################################################

 ddd_w_brand_dev = ddd_w_deviation_units[['brd_nm', 'deviation']].copy()
 #ddd_w_brand_dev

 ddd_w_brand_dev.to_excel(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\ddd_w_brand_deviation.xlsx" %(prid),index=False)

 ddd_w_brand_deviation = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\brand_deviation.xlsx" %(prid))
 ddd_w_brand_deviation

 ddd_w_brand_res = ddd_w_brand_deviation.groupby('brd_nm')
 #ddd_w_brand_res

 ddd_w_brand_out = ddd_w_brand_res.mean()
 #ddd_w_brand_out

 ddd_w_brand_out.to_excel(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\ddd_w_brand_dev_output.xlsx" %(prid))

 ddd_w_brand_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\ddd_w_brand_dev_output.xlsx" %(prid))

 sns.set(rc={'figure.figsize':(11.7,8.27)})
 ddd_w_brand_sns_plot = sns.barplot(x="deviation",y="brd_nm",data=ddd_w_brand_df)
 ddd_w_brand_sns_plot.figure.savefig(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\ddd_w_brand_output.pdf" %(prid),bbox_inches="tight",pad_inches=1)

###############################################PRODUCT##########################################################################

 ddd_w_product_dev = ddd_w_deviation_units[['prod_nm', 'deviation']].copy()
 #ddd_w_product_dev

 ddd_w_product_dev.to_excel(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\ddd_w_product_deviation.xlsx" %(prid),index=False)

 ddd_w_product_deviation = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\ddd_w_product_deviation.xlsx" %(prid))
 ddd_w_product_deviation

 ddd_w_product_res = ddd_w_product_deviation.groupby('prod_nm')
 #ddd_w_product_res

 ddd_w_product_out = ddd_w_product_res.mean()
 #ddd_w_product_out

 ddd_w_product_out.to_excel(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\ddd_w_product_dev_output.xlsx" %(prid))

 ddd_w_product_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\ddd_w_product_dev_output.xlsx" %(prid))

 sns.set(rc={'figure.figsize':(11.7,8.27)})
 ddd_w_prod_sns_plot = sns.barplot(x="deviation",y="prod_nm",data=ddd_w_product_df)
 ddd_w_prod_sns_plot.figure.savefig(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\ddd_w_product_output.pdf" %(prid),bbox_inches="tight",pad_inches=1)

#################################Moving the image files to output folder######################################################

if ddd_w_units_deviations_flag == 'Y':
  pass
else:
 sourcepath = r'C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\\' %(prid) 
 sourcefiles = os.listdir(sourcepath)
 destinationpath = r'C:\Users\%s\Desktop\unaligned\ddd_weekly\output\\' %(prid)
 for file in sourcefiles:
    if file.endswith('.pdf %(prid)'):
        shutil.move(os.path.join(sourcepath,file), os.path.join(destinationpath,file))

###################################################Deleting unwanted files after run from temp folder########################################################

import os
import shutil

for root, dirs, files in os.walk(r'C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\\' %(prid)):
    for f in files:
        os.unlink(os.path.join(root, f))
    for d in dirs:
        shutil.rmtree(os.path.join(root, d))
       
#################################################DDD_MONTHLY#######################################################################

#########################UNITS##################################################################################################################

ddd_m_units = input_ddd_m_df.pivot_table(index = ["mkt_nm","brd_nm","prod_nm"],values = ["tot_units_new","tot_units_old"])
ddd_m_units["deviation"] = ((ddd_m_units.tot_units_new - ddd_m_units.tot_units_old)/ddd_m_units.tot_units_old)*100
#ddd_m_units
ddd_m_units["deviation"] = ddd_m_units["deviation"].round(round(1))
#ddd_m_units
ddd_m_units_deviation = ddd_m_units[((ddd_m_units['deviation'] <= -5) | (ddd_m_units['deviation'] >=5))]
#ddd_m_units_deviation
ddd_m_units_deviations = ddd_m_units_deviation[((ddd_m_units_deviation['tot_units_new'] != 0) & (ddd_m_units_deviation['tot_units_old'] != 0))]
#ddd_m_units_deviations
ddd_m_units_zero_sales = ddd_m_units[(ddd_m_units['tot_units_new'] == 0) | (ddd_m_units['tot_units_old'] == 0)]
#ddd_m_units_zero_sales

ddd_m_units.to_excel(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\temp\ddd_m_units.xlsx" %(prid),sheet_name = "ddd_m_units")

ddd_m_units_deviations_flag = 'N'
ddd_m_units_zero_sales_flag = 'N'
ddd_m_dollars_deviations_flag = 'N'
ddd_m_dollars_zero_sales_flag = 'N'

if ddd_m_units_deviations.empty == True:
 ddd_m_units_deviations_flag = 'Y'
else:
 ddd_m_units_deviations.to_excel(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\temp\deviations_units.xlsx" %(prid),sheet_name = "deviations_units")

if ddd_m_units_zero_sales.empty == True:
 ddd_m_units_zero_sales_flag = 'Y'
else:
 ddd_m_units_zero_sales.to_excel(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\temp\zero_sales_units.xlsx" %(prid),sheet_name = "zero_sales_units")

##Zero_sales_units:

if ddd_m_units_deviations_flag == 'Y':
  pass
else:
 zs_ddd_m_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\temp\zero_sales_units.xlsx" %(prid))
 ddd_m_zero_sales_units = zs_ddd_m_df.fillna(method = "ffill")
#ddd_m_zero_sales_units

 ddd_m_zero_sales_units.loc[(ddd_m_zero_sales_units['tot_units_old'] == 0) & (ddd_m_zero_sales_units['tot_units_new'] == 0), 'comments'] = "no comments"
 ddd_m_zero_sales_units.loc[(ddd_m_zero_sales_units['tot_units_new'] == 0) & (ddd_m_zero_sales_units['tot_units_old'] != 0), 'comments'] = ddd_m_zero_sales_units['prod_nm'] + " sales dropped to zero under " + ddd_m_zero_sales_units["mkt_nm"]
 ddd_m_zero_sales_units.loc[(ddd_m_zero_sales_units['tot_units_old'] == 0) & (ddd_m_zero_sales_units['tot_units_new'] != 0), 'comments'] = ddd_m_zero_sales_units['prod_nm'] + " is flowing under " + ddd_m_zero_sales_units["mkt_nm"]

 ddd_m_zero_sales_units = ddd_m_zero_sales_units.sort_values(by='prod_nm')
 
##Deviation_units:

if ddd_m_units_zero_sales_flag == 'Y':
  pass
else:
 dev_ddd_m_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\temp\deviations_units.xlsx" %(prid))
 ddd_m_deviation_units = dev_ddd_m_df.fillna(method = "ffill")
#ddd_m_deviation_units
 
 ddd_m_deviation_units.loc[ddd_m_deviation_units['deviation'] > 5, 'comments'] = ddd_m_deviation_units['brd_nm'] + " sales is increased by " + ddd_m_deviation_units['deviation'].map(str) + "%" + " in " + ddd_m_deviation_units['mkt_nm']
 ddd_m_deviation_units.loc[ddd_m_deviation_units['deviation'] < -5, 'comments'] = ddd_m_deviation_units['brd_nm'] + " sales is dropped by " + ddd_m_deviation_units['deviation'].map(str) + "%" + " in " + ddd_m_deviation_units['mkt_nm']
 
 ddd_m_deviation_units = ddd_m_deviation_units.sort_values(by='prod_nm')

#############################################DOLLARS##############################################################################################

ddd_m_dollars = input_ddd_m_df.pivot_table(index = ["mkt_nm","brd_nm","prod_nm"],values = ["tot_dolrs_new","tot_dolrs_old"])
ddd_m_dollars["deviation"] = ((ddd_m_dollars.tot_dolrs_new - ddd_m_dollars.tot_dolrs_old)/ddd_m_dollars.tot_dolrs_old)*100
#ddd_m_dollars
ddd_m_dollars["deviation"] = ddd_m_dollars["deviation"].round(round(1))
#ddd_m_dollars
ddd_m_dollars_deviation = ddd_m_dollars[((ddd_m_dollars['deviation'] <= -5) | (ddd_m_dollars['deviation'] >=5))]
#ddd_m_dollars_deviation
ddd_m_dollars_deviations = ddd_m_dollars_deviation[((ddd_m_dollars_deviation['tot_dolrs_new'] != 0) & (ddd_m_dollars_deviation['tot_dolrs_old'] != 0))]
#ddd_m_dollars_deviations
ddd_m_dollars_zero_sales = ddd_m_dollars[(ddd_m_dollars['tot_dolrs_new'] == 0) | (ddd_m_dollars['tot_dolrs_old'] == 0)]
#ddd_m_dollars_zero_sales

ddd_m_dollars.to_excel(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\temp\ddd_m_dollars.xlsx" %(prid),sheet_name = "ddd_m_dollars")

if ddd_m_dollars_deviations.empty == True:
 ddd_m_dollars_deviations_flag = 'Y'
else:
 ddd_m_dollars_deviations.to_excel(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\temp\deviations_dollars.xlsx" %(prid),sheet_name = "deviations_dollars")
 
if ddd_m_dollars_zero_sales.empty == True:
 ddd_m_dollars_zero_sales_flag = 'Y'
else:
 ddd_m_dollars_zero_sales.to_excel(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\temp\zero_sales_dollars.xlsx" %(prid),sheet_name = "zero_sales_dollars")

##Zero_sales_dollars:

if ddd_m_dollars_deviations_flag == 'Y':
  pass
else:
 Zero_sales_dollars_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\temp\zero_sales_dollars.xlsx" %(prid))
 ddd_m_zero_sales_dollars = Zero_sales_dollars_df.fillna(method = "ffill")
#ddd_m_zero_sales_dollars

 ddd_m_zero_sales_dollars.loc[(ddd_m_zero_sales_dollars['tot_dolrs_old'] == 0) & (ddd_m_zero_sales_dollars['tot_dolrs_new'] == 0), 'comments'] = "no comments"
 ddd_m_zero_sales_dollars.loc[(ddd_m_zero_sales_dollars['tot_dolrs_new'] == 0) & (ddd_m_zero_sales_dollars['tot_dolrs_old'] != 0), 'comments'] = ddd_m_zero_sales_dollars['prod_nm'] + " sales dropped to zero under " + ddd_m_zero_sales_dollars["mkt_nm"]
 ddd_m_zero_sales_dollars.loc[(ddd_m_zero_sales_dollars['tot_dolrs_old'] == 0) & (ddd_m_zero_sales_dollars['tot_dolrs_new'] != 0), 'comments'] = ddd_m_zero_sales_dollars['prod_nm'] + " is flowing under " + ddd_m_zero_sales_dollars["mkt_nm"]
 
 ddd_m_zero_sales_dollars = ddd_m_zero_sales_dollars.sort_values(by='prod_nm')

##Deviation_dollars:

if ddd_m_dollars_zero_sales_flag == 'Y':
  pass
else:
 Deviation_dollars_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\temp\deviations_dollars.xlsx" %(prid))
 ddd_m_deviation_dollars = Deviation_dollars_df.fillna(method = "ffill")
#ddd_m_deviation_dollars

#deviation_dollars['deviation']= deviation_dollars['deviation'].astype(str)

 ddd_m_deviation_dollars.loc[ddd_m_deviation_dollars['deviation'] > 5, 'comments'] = ddd_m_deviation_dollars['brd_nm'] + " sales is increased by " + ddd_m_deviation_dollars['deviation'].map(str) + "%" + " in " + ddd_m_deviation_dollars['mkt_nm']
 ddd_m_deviation_dollars.loc[ddd_m_deviation_dollars['deviation'] < -5, 'comments'] = ddd_m_deviation_dollars['brd_nm'] + " sales is dropped by " + ddd_m_deviation_dollars['deviation'].map(str) + "%" + " in " + ddd_m_deviation_dollars['mkt_nm']
 ddd_m_deviation_dollars = ddd_m_deviation_dollars.sort_values(by='prod_nm')
 
 #############################Writing output to excel file##################################################################################

if [ddd_m_units_deviations_flag == 'Y'] or [ddd_m_units_zero_sales_flag == 'Y'] or [ddd_m_dollars_deviations_flag == 'Y'] or [ddd_m_dollars_zero_sales_flag == 'Y']:
    with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\output\ddd_m_output.xlsx" %(prid)) as writer:    
      ddd_m_units.to_excel(writer,sheet_name = "ddd_m_units")
      ddd_m_dollars.to_excel(writer,sheet_name = "ddd_m_dollars")
    
    if ddd_m_units_deviations_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\output\ddd_m_output.xlsx" %(prid))
        book.create_sheet("deviation_units")
        sheet = book["deviation_units"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\output\ddd_m_output.xlsx" %(prid))
        
    else:
        with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\output\ddd_m_output.xlsx" %(prid), engine='openpyxl', mode='a') as writer: 
         ddd_m_deviation_units.to_excel(writer,sheet_name = "deviation_units")

    if ddd_m_units_zero_sales_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\output\ddd_m_output.xlsx" %(prid))
        book.create_sheet("Zero_sales_units")
        sheet = book["Zero_sales_units"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\output\ddd_m_output.xlsx" %(prid))
        
    else:
        with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\output\ddd_m_output.xlsx" %(prid), engine='openpyxl', mode='a') as writer: 
         ddd_m_zero_sales_units.to_excel(writer,sheet_name = "Zero_sales_units")
   
    if ddd_m_dollars_deviations_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\output\ddd_m_output.xlsx" %(prid))
        book.create_sheet("deviation_dollars")
        sheet = book["deviation_dollars"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\output\ddd_m_output.xlsx" %(prid))
        
    else:
        with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\output\ddd_m_output.xlsx" %(prid), engine='openpyxl', mode='a') as writer: 
         ddd_m_deviation_dollars.to_excel(writer,sheet_name = "deviation_dollars")
  

    if ddd_m_dollars_zero_sales_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\output\ddd_m_output.xlsx" %(prid))
        book.create_sheet("Zero_sales_dollars")
        sheet = book["Zero_sales_dollars"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\output\ddd_m_output.xlsx" %(prid))
        
    else:
        with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\output\ddd_m_output.xlsx" %(prid), engine='openpyxl', mode='a') as writer: 
         ddd_m_zero_sales_dollars.to_excel(writer,sheet_name = "Zero_sales_dollars")

##################################################DDD_Monthly_Data_Visualization#########################################################################

####################################################MARKET##################################################################################

if ddd_m_units_deviations_flag == 'Y':
  pass
else:

 ddd_m_mkt_dev = ddd_m_deviation_units[['mkt_nm', 'deviation']].copy()
 #ddd_m_mkt_dev

 ddd_m_mkt_dev.to_excel(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\temp\ddd_m_mkt_deviation.xlsx" %(prid),index=False)
  
 ddd_m_market_deviation = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\temp\ddd_m_mkt_deviation.xlsx" %(prid))
 ddd_m_market_deviation

 ddd_m_mar_res = ddd_m_market_deviation.groupby('mkt_nm')
 #ddd_m_mar_res

 ddd_m_mar_out = ddd_m_mar_res.mean()
 #ddd_m_mar_out

 ddd_m_mar_out.to_excel(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\temp\ddd_m_mkt_dev_output.xlsx" %(prid))

 ddd_m_mar_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\temp\ddd_m_mkt_dev_output.xlsx" %(prid))

 sns.set(rc={'figure.figsize':(11.7,8.27)})
 ddd_m_mar_sns_plot = sns.barplot(x="deviation",y="mkt_nm",data=ddd_m_mar_df)
 ddd_m_mar_sns_plot.figure.savefig(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\temp\ddd_m_mar_output.pdf" %(prid),bbox_inches="tight",pad_inches=1)

#######################################BRAND######################################################################################

 ddd_m_brand_dev = ddd_m_deviation_units[['brd_nm', 'deviation']].copy()
 #ddd_m_brand_dev

 ddd_m_brand_dev.to_excel(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\temp\ddd_m_brand_deviation.xlsx" %(prid),index=False)

 ddd_m_brand_deviation = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\ddd_weekly\temp\brand_deviation.xlsx" %(prid))
 ddd_m_brand_deviation

 ddd_m_brand_res = ddd_m_brand_deviation.groupby('brd_nm')
 #ddd_m_brand_res

 ddd_m_brand_out = ddd_m_brand_res.mean()
 #ddd_m_brand_out

 ddd_m_brand_out.to_excel(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\temp\ddd_m_brand_dev_output.xlsx" %(prid))

 ddd_m_brand_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\temp\ddd_m_brand_dev_output.xlsx" %(prid))

 sns.set(rc={'figure.figsize':(11.7,8.27)})
 ddd_m_brand_sns_plot = sns.barplot(x="deviation",y="brd_nm",data=ddd_m_brand_df)
 ddd_m_brand_sns_plot.figure.savefig(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\temp\ddd_m_brand_output.pdf" %(prid),bbox_inches="tight",pad_inches=1)

###############################################PRODUCT##########################################################################

 ddd_m_product_dev = ddd_m_deviation_units[['prod_nm', 'deviation']].copy()
 #ddd_m_product_dev

 ddd_m_product_dev.to_excel(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\temp\ddd_m_product_deviation.xlsx" %(prid),index=False)

 ddd_m_product_deviation = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\temp\ddd_m_product_deviation.xlsx" %(prid))
 ddd_m_product_deviation

 ddd_m_product_res = ddd_m_product_deviation.groupby('prod_nm')
 #ddd_m_product_res

 ddd_m_product_out = ddd_m_product_res.mean()
 #ddd_m_product_out

 ddd_m_product_out.to_excel(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\temp\ddd_m_product_dev_output.xlsx" %(prid))

 ddd_m_product_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\temp\ddd_m_product_dev_output.xlsx" %(prid))

 sns.set(rc={'figure.figsize':(11.7,8.27)})
 ddd_m_prod_sns_plot = sns.barplot(x="deviation",y="prod_nm",data=ddd_m_product_df)
 ddd_m_prod_sns_plot.figure.savefig(r"C:\Users\%s\Desktop\unaligned\ddd_monthly\temp\ddd_m_product_output.pdf" %(prid),bbox_inches="tight",pad_inches=1)

#################################Moving the image files to output folder######################################################

if ddd_m_units_deviations_flag == 'Y':
  pass
else:
 sourcepath = r'C:\Users\%s\Desktop\unaligned\ddd_monthly\temp\\' %(prid)
 sourcefiles = os.listdir(sourcepath)
 destinationpath = r'C:\Users\%s\Desktop\unaligned\ddd_monthly\output\\' %(prid)
 for file in sourcefiles:
    if file.endswith('.pdf %(prid)'):
        shutil.move(os.path.join(sourcepath,file), os.path.join(destinationpath,file))

###################################################Deleting unwanted files after run from temp folder########################################################

import os
import shutil

for root, dirs, files in os.walk(r'C:\Users\%s\Desktop\unaligned\ddd_monthly\temp\\' %(prid)):
    for f in files:
        os.unlink(os.path.join(root, f))
    for d in dirs:
        shutil.rmtree(os.path.join(root, d))

##########################################XPO WEEKLY############################################################################################


#########################Nrx##################################################################################################################

xpo_w_nrx = input_xpo_w_df.pivot_table(index = ["mkt_nm","brd_nm","prod_nm"],values = ["nrx_new","nrx_old"])
xpo_w_nrx["deviation"] = ((xpo_w_nrx.nrx_new - xpo_w_nrx.nrx_old)/xpo_w_nrx.nrx_old)*100
#xpo_w_nrx
xpo_w_nrx["deviation"] = xpo_w_nrx["deviation"].round(round(1))
#xpo_w_nrx
xpo_w_nrx_deviation = xpo_w_nrx[((xpo_w_nrx['deviation'] <= -5) | (xpo_w_nrx['deviation'] >=5))]
#xpo_w_nrx_deviation
xpo_w_nrx_deviations = xpo_w_nrx_deviation[((xpo_w_nrx_deviation['nrx_new'] != 0) & (xpo_w_nrx_deviation['nrx_old'] != 0))]
#xpo_w_nrx_deviations
xpo_w_nrx_zero_sales = xpo_w_nrx[(xpo_w_nrx['nrx_new'] == 0) | (xpo_w_nrx['nrx_old'] == 0)]
#xpo_w_nrx_zero_sales

xpo_w_nrx.to_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\xpo_w_nrx.xlsx" %(prid),sheet_name = "xpo_w_nrx")

xpo_w_nrx_deviations_flag = 'N'
xpo_w_nrx_zero_sales_flag = 'N'
xpo_w_trx_deviations_flag = 'N'
xpo_w_trx_zero_sales_flag = 'N'

if xpo_w_nrx_deviations.empty == True:
 xpo_w_nrx_deviations_flag = 'Y'
else:
 xpo_w_nrx_deviations.to_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\deviations_nrx.xlsx" %(prid),sheet_name = "deviations_nrx")

if xpo_w_nrx_zero_sales.empty == True:
 xpo_w_nrx_zero_sales_flag = 'Y'
else:
 xpo_w_nrx_zero_sales.to_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\zero_sales_nrx.xlsx" %(prid),sheet_name = "zero_sales_nrx")

##Zero_sales_nrx:

if xpo_w_nrx_zero_sales_flag == 'Y':
  pass
else:
 zs_xpo_w_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\zero_sales_nrx.xlsx" %(prid))
 xpo_w_zero_sales_nrx = zs_xpo_w_df.fillna(method = "ffill")
#xpo_w_zero_sales_nrx

 xpo_w_zero_sales_nrx.loc[(xpo_w_zero_sales_nrx['nrx_old'] == 0) & (xpo_w_zero_sales_nrx['nrx_new'] == 0), 'comments'] = "no comments"
 xpo_w_zero_sales_nrx.loc[(xpo_w_zero_sales_nrx['nrx_new'] == 0) & (xpo_w_zero_sales_nrx['nrx_old'] != 0), 'comments'] = xpo_w_zero_sales_nrx['prod_nm'] + " sales dropped to zero under " + xpo_w_zero_sales_nrx["mkt_nm"]
 xpo_w_zero_sales_nrx.loc[(xpo_w_zero_sales_nrx['nrx_old'] == 0) & (xpo_w_zero_sales_nrx['nrx_new'] != 0), 'comments'] = xpo_w_zero_sales_nrx['prod_nm'] + " is flowing under " + xpo_w_zero_sales_nrx["mkt_nm"]
 
 xpo_w_zero_sales_nrx = xpo_w_zero_sales_nrx.sort_values(by='prod_nm')

##Deviation_nrx:

if xpo_w_nrx_deviations_flag == 'Y':
  pass
else:
 dev_xpo_w_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\deviations_nrx.xlsx" %(prid))
 xpo_w_deviation_nrx = dev_xpo_w_df.fillna(method = "ffill")
#xpo_w_deviation_nrx

 xpo_w_deviation_nrx.loc[xpo_w_deviation_nrx['deviation'] > 5, 'comments'] = xpo_w_deviation_nrx['brd_nm'] + " sales is increased by " + xpo_w_deviation_nrx['deviation'].map(str) + "%" + " in " + xpo_w_deviation_nrx['mkt_nm']
 xpo_w_deviation_nrx.loc[xpo_w_deviation_nrx['deviation'] < -5, 'comments'] = xpo_w_deviation_nrx['brd_nm'] + " sales is dropped by " + xpo_w_deviation_nrx['deviation'].map(str) + "%" + " in " + xpo_w_deviation_nrx['mkt_nm']
 
 xpo_w_deviation_nrx = xpo_w_deviation_nrx.sort_values(by='prod_nm')

#############################################Trx##############################################################################################

xpo_w_trx = input_xpo_w_df.pivot_table(index = ["mkt_nm","brd_nm","prod_nm"],values = ["trx_new","trx_old"])
xpo_w_trx["deviation"] = ((xpo_w_trx.trx_new - xpo_w_trx.trx_old)/xpo_w_trx.trx_old)*100
#xpo_w_trx
xpo_w_trx["deviation"] = xpo_w_trx["deviation"].round(round(1))
#xpo_w_trx
xpo_w_trx_deviation = xpo_w_trx[((xpo_w_trx['deviation'] <= -5) | (xpo_w_trx['deviation'] >=5))]
#xpo_w_trx_deviation
xpo_w_trx_deviations = xpo_w_trx_deviation[((xpo_w_trx_deviation['trx_new'] != 0) & (xpo_w_trx_deviation['trx_old'] != 0))]
#xpo_w_trx_deviations
xpo_w_trx_zero_sales = xpo_w_trx[(xpo_w_trx['trx_new'] == 0) | (xpo_w_trx['trx_old'] == 0)]
#xpo_w_trx_zero_sales

xpo_w_trx.to_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\xpo_w_trx.xlsx" %(prid),sheet_name = "xpo_w_trx")

if xpo_w_trx_deviations.empty == True:
 xpo_w_trx_deviations_flag = 'Y'
else:
 xpo_w_trx_deviations.to_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\deviations_trx.xlsx" %(prid),sheet_name = "deviations_trx")
 
if xpo_w_trx_zero_sales.empty == True:
 xpo_w_trx_zero_sales_flag = 'Y'
else:
 xpo_w_trx_zero_sales.to_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\zero_sales_trx.xlsx" %(prid),sheet_name = "zero_sales_trx")

##Zero_sales_trx:

if xpo_w_trx_zero_sales_flag == 'Y':
  pass
else:
 Zero_sales_trx_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\zero_sales_trx.xlsx" %(prid))
 xpo_w_zero_sales_trx = Zero_sales_trx_df.fillna(method = "ffill")
#zero_sales_trx

 xpo_w_zero_sales_trx.loc[(xpo_w_zero_sales_trx['trx_old'] == 0) & (xpo_w_zero_sales_trx['trx_new'] == 0), 'comments'] = "no comments"
 xpo_w_zero_sales_trx.loc[(xpo_w_zero_sales_trx['trx_new'] == 0) & (xpo_w_zero_sales_trx['trx_old'] != 0), 'comments'] = xpo_w_zero_sales_trx['prod_nm'] + " sales dropped to zero under " + xpo_w_zero_sales_trx["mkt_nm"]
 xpo_w_zero_sales_trx.loc[(xpo_w_zero_sales_trx['trx_old'] == 0) & (xpo_w_zero_sales_trx['trx_new'] != 0), 'comments'] = xpo_w_zero_sales_trx['prod_nm'] + " is flowing under " + xpo_w_zero_sales_trx["mkt_nm"]
 
 xpo_w_zero_sales_trx = xpo_w_zero_sales_trx.sort_values(by='prod_nm')

##Deviation_trx:

if xpo_w_trx_deviations_flag == 'Y':
  pass
else:
 Deviation_trx_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\deviations_trx.xlsx" %(prid))
 xpo_w_deviation_trx = Deviation_trx_df.fillna(method = "ffill")
#xpo_w_deviation_trx

#deviation_trx['deviation']= deviation_trx['deviation'].astype(str)

 xpo_w_deviation_trx.loc[xpo_w_deviation_trx['deviation'] > 5, 'comments'] = xpo_w_deviation_trx['brd_nm'] + " sales is increased by " + xpo_w_deviation_trx['deviation'].map(str) + "%" + " in " + xpo_w_deviation_trx['mkt_nm']
 xpo_w_deviation_trx.loc[xpo_w_deviation_trx['deviation'] < -5, 'comments'] = xpo_w_deviation_trx['brd_nm'] + " sales is dropped by " + xpo_w_deviation_trx['deviation'].map(str) + "%" + " in " + xpo_w_deviation_trx['mkt_nm']
 xpo_w_deviation_trx = xpo_w_deviation_trx.sort_values(by='prod_nm')
 
 ####################################Wrting the output to excel file################################################################### 	

if [xpo_w_nrx_deviations_flag == 'Y'] or [xpo_w_nrx_zero_sales_flag == 'Y'] or [xpo_w_trx_deviations_flag == 'Y'] or [xpo_w_trx_zero_sales_flag == 'Y']:
    with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx" %(prid)) as writer:    
      xpo_w_nrx.to_excel(writer,sheet_name = "xpo_w_nrx")
      xpo_w_trx.to_excel(writer,sheet_name = "xpo_w_trx")
    
    if xpo_w_nrx_deviations_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx" %(prid))
        book.create_sheet("deviation_nrx")
        sheet = book["deviation_nrx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx" %(prid))
        
    else:
        with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx" %(prid), engine='openpyxl', mode='a') as writer: 
         xpo_w_deviation_nrx.to_excel(writer,sheet_name = "deviation_nrx")

    if xpo_w_nrx_zero_sales_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx" %(prid))
        book.create_sheet("Zero_sales_nrx")
        sheet = book["Zero_sales_nrx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx" %(prid))
        
    else:
        with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx" %(prid), engine='openpyxl', mode='a') as writer: 
         xpo_w_zero_sales_nrx.to_excel(writer,sheet_name = "zero_sales_nrx")
   
    if xpo_w_trx_deviations_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx" %(prid))
        book.create_sheet("deviation_trx")
        sheet = book["deviation_trx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx" %(prid))
        
    else:
        with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx" %(prid), engine='openpyxl', mode='a') as writer: 
         xpo_w_deviation_trx.to_excel(writer,sheet_name = "deviation_trx")
  

    if xpo_w_trx_zero_sales_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx" %(prid))
        book.create_sheet("Zero_sales_trx")
        sheet = book["Zero_sales_trx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx" %(prid))
        
    else:
        with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx" %(prid), engine='openpyxl', mode='a') as writer: 
         xpo_w_zero_sales_trx.to_excel(writer,sheet_name = "Zero_sales_trx")

##################################################XPO_WEEKLY_Data_Visualization#########################################################################

####################################################MARKET##################################################################################

if xpo_w_nrx_deviations_flag == 'Y':
  pass
else:
 xpo_w_mkt_dev = xpo_w_deviation_nrx[['mkt_nm', 'deviation']].copy()
 #xpo_w_mkt_dev

 xpo_w_mkt_dev.to_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\xpo_w_mkt_deviation.xlsx" %(prid),index=False)

 xpo_w_market_deviation = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\xpo_w_mkt_deviation.xlsx" %(prid))
 xpo_w_market_deviation

 xpo_w_mar_res = xpo_w_market_deviation.groupby('mkt_nm')
 #xpo_w_mar_res

 xpo_w_mar_out = xpo_w_mar_res.mean()
 #xpo_w_mar_out

 xpo_w_mar_out.to_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\xpo_w_mkt_dev_output.xlsx" %(prid))

 xpo_w_mar_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\xpo_w_mkt_dev_output.xlsx" %(prid))
  
 sns.set(rc={'figure.figsize':(11.7,8.27)})
 xpo_w_mar_sns_plot = sns.barplot(x="deviation",y="mkt_nm",data=xpo_w_mar_df)
 xpo_w_mar_sns_plot.figure.savefig(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\xpo_w_mar_output.pdf" %(prid),bbox_inches="tight",pad_inches=1)

#######################################BRAND######################################################################################

 xpo_w_brand_dev = xpo_w_deviation_nrx[['brd_nm', 'deviation']].copy()
 #xpo_w_brand_dev

 xpo_w_brand_dev.to_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\xpo_w_brand_deviation.xlsx" %(prid),index=False)

 xpo_w_brand_deviation = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\xpo_w_brand_deviation.xlsx" %(prid))
 xpo_w_brand_deviation

 xpo_w_brand_res = xpo_w_brand_deviation.groupby('brd_nm')
 #xpo_w_brand_res

 xpo_w_brand_out = xpo_w_brand_res.mean()
 #xpo_w_brand_out

 xpo_w_brand_out.to_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\xpo_w_brand_dev_output.xlsx" %(prid))

 xpo_w_brand_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\xpo_w_brand_dev_output.xlsx" %(prid))

 sns.set(rc={'figure.figsize':(11.7,8.27)})
 xpo_w_brand_sns_plot = sns.barplot(x="deviation",y="brd_nm",data=xpo_w_brand_df)
 xpo_w_brand_sns_plot.figure.savefig(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\xpo_w_brand_output.pdf" %(prid),bbox_inches="tight",pad_inches=1)

###############################################PRODUCT##########################################################################

 xpo_w_product_dev = xpo_w_deviation_nrx[['prod_nm', 'deviation']].copy()
 #xpo_w_product_dev

 xpo_w_product_dev.to_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\xpo_w_product_deviation.xlsx" %(prid),index=False)

 xpo_w_product_deviation = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\xpo_w_product_deviation.xlsx" %(prid))
 xpo_w_product_deviation

 xpo_w_product_res = xpo_w_product_deviation.groupby('prod_nm')
 #xpo_w_product_res

 xpo_w_product_out = xpo_w_product_res.mean()
 #xpo_w_product_out

 xpo_w_product_out.to_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\xpo_w_product_dev_output.xlsx" %(prid))

 xpo_w_product_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\xpo_w_product_dev_output.xlsx" %(prid))

 sns.set(rc={'figure.figsize':(11.7,8.27)})
 xpo_w_prod_sns_plot = sns.barplot(x="deviation",y="prod_nm",data=xpo_w_product_df)
 xpo_w_prod_sns_plot.figure.savefig(r"C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\xpo_w_product_output.pdf" %(prid),bbox_inches="tight",pad_inches=1)

#################################Moving the image files to output folder######################################################

if xpo_w_nrx_deviations_flag == 'Y':
  pass
else:
 sourcepath = r'C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\\' %(prid)
 sourcefiles = os.listdir(sourcepath)
 destinationpath = r'C:\Users\%s\Desktop\unaligned\xpo_weekly\output\\' %(prid)
 for file in sourcefiles:
    if file.endswith('.pdf'):
        shutil.move(os.path.join(sourcepath,file), os.path.join(destinationpath,file))

###################################################Deleting unwanted files after run from temp folder########################################################

import os
import shutil

for root, dirs, files in os.walk(r'C:\Users\%s\Desktop\unaligned\xpo_weekly\temp\\' %(prid)):
    for f in files:
        os.unlink(os.path.join(root, f))
    for d in dirs:
        shutil.rmtree(os.path.join(root, d))
        
#####################################################XPOPD_WEEKLY################################################################


#########################Nrx##################################################################################################################

xpopd_w_nrx = input_xpopd_w_df.pivot_table(index = ["mkt_nm","brd_nm","prod_nm"],values = ["nrx_new","nrx_old"])
xpopd_w_nrx["deviation"] = ((xpopd_w_nrx.nrx_new - xpopd_w_nrx.nrx_old)/xpopd_w_nrx.nrx_old)*100
#xpopd_w_nrx
xpopd_w_nrx["deviation"] = xpopd_w_nrx["deviation"].round(round(1))
#xpopd_w_nrx
xpopd_w_nrx_deviation = xpopd_w_nrx[((xpopd_w_nrx['deviation'] <= -5) | (xpopd_w_nrx['deviation'] >=5))]
#xpopd_w_nrx_deviation
xpopd_w_nrx_deviations = xpopd_w_nrx_deviation[((xpopd_w_nrx_deviation['nrx_new'] != 0) & (xpopd_w_nrx_deviation['nrx_old'] != 0))]
#xpopd_w_nrx_deviations
xpopd_w_nrx_zero_sales = xpopd_w_nrx[(xpopd_w_nrx['nrx_new'] == 0) | (xpopd_w_nrx['nrx_old'] == 0)]
#xpopd_w_nrx_zero_sales

xpopd_w_nrx.to_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\temp\xpopd_w_nrx.xlsx" %(prid),sheet_name = "xpopd_w_nrx")

xpopd_w_nrx_deviations_flag = 'N'
xpopd_w_nrx_zero_sales_flag = 'N'
xpopd_w_trx_deviations_flag = 'N'
xpopd_w_trx_zero_sales_flag = 'N'

if xpopd_w_nrx_deviations.empty == True:
 xpopd_w_nrx_deviations_flag = 'Y'
else:
 xpopd_w_nrx_deviations.to_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\temp\deviations_nrx.xlsx" %(prid),sheet_name = "deviations_nrx")

if xpopd_w_nrx_zero_sales.empty == True:
 xpopd_w_nrx_zero_sales_flag = 'Y'
else:
 xpopd_w_nrx_zero_sales.to_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\temp\zero_sales_nrx.xlsx" %(prid),sheet_name = "zero_sales_nrx")

##Zero_sales_nrx:

if xpopd_w_nrx_zero_sales_flag == 'Y':
  pass
else:
 zs_xpopd_w_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\temp\zero_sales_nrx.xlsx" %(prid))
 xpopd_w_zero_sales_nrx = zs_xpopd_w_df.fillna(method = "ffill")
#xpopd_w_zero_sales_nrx

 xpopd_w_zero_sales_nrx.loc[(xpopd_w_zero_sales_nrx['nrx_old'] == 0) & (xpopd_w_zero_sales_nrx['nrx_new'] == 0), 'comments'] = "no comments"
 xpopd_w_zero_sales_nrx.loc[(xpopd_w_zero_sales_nrx['nrx_new'] == 0) & (xpopd_w_zero_sales_nrx['nrx_old'] != 0), 'comments'] = xpopd_w_zero_sales_nrx['prod_nm'] + " sales dropped to zero under " + xpopd_w_zero_sales_nrx["mkt_nm"]
 xpopd_w_zero_sales_nrx.loc[(xpopd_w_zero_sales_nrx['nrx_old'] == 0) & (xpopd_w_zero_sales_nrx['nrx_new'] != 0), 'comments'] = xpopd_w_zero_sales_nrx['prod_nm'] + " is flowing under " + xpopd_w_zero_sales_nrx["mkt_nm"]
 
 xpopd_w_zero_sales_nrx = xpopd_w_zero_sales_nrx.sort_values(by='prod_nm')

##Deviation_nrx:

if xpopd_w_nrx_deviations_flag == 'Y':
  pass
else:
 dev_xpopd_w_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\temp\deviations_nrx.xlsx" %(prid))
 xpopd_w_deviation_nrx = dev_xpopd_w_df.fillna(method = "ffill")
#xpopd_w_deviation_nrx

#deviation_nrx['deviation']= deviation_nrx['deviation'].astype(str)
 
 xpopd_w_deviation_nrx.loc[xpopd_w_deviation_nrx['deviation'] > 5, 'comments'] = xpopd_w_deviation_nrx['brd_nm'] + " sales is increased by " + xpopd_w_deviation_nrx['deviation'].map(str) + "%" + " in " + xpopd_w_deviation_nrx['mkt_nm']
 xpopd_w_deviation_nrx.loc[xpopd_w_deviation_nrx['deviation'] < -5, 'comments'] = xpopd_w_deviation_nrx['brd_nm'] + " sales is dropped by " + xpopd_w_deviation_nrx['deviation'].map(str) + "%" + " in " + xpopd_w_deviation_nrx['mkt_nm']
 
 xpopd_w_deviation_nrx = xpopd_w_deviation_nrx.sort_values(by='prod_nm')

#############################################Trx##############################################################################################

xpopd_w_trx = input_xpopd_w_df.pivot_table(index = ["mkt_nm","brd_nm","prod_nm"],values = ["trx_new","trx_old"])
xpopd_w_trx["deviation"] = ((xpopd_w_trx.trx_new - xpopd_w_trx.trx_old)/xpopd_w_trx.trx_old)*100
#xpopd_w_trx
xpopd_w_trx["deviation"] = xpopd_w_trx["deviation"].round(round(1))
#xpopd_w_trx
xpopd_w_trx_deviation = xpopd_w_trx[((xpopd_w_trx['deviation'] <= -5) | (xpopd_w_trx['deviation'] >=5))]
#xpopd_w_trx_deviation
xpopd_w_trx_deviations = xpopd_w_trx_deviation[((xpopd_w_trx_deviation['trx_new'] != 0) & (xpopd_w_trx_deviation['trx_old'] != 0))]
#xpopd_w_trx_deviations
xpopd_w_trx_zero_sales = xpopd_w_trx[(xpopd_w_trx['trx_new'] == 0) | (xpopd_w_trx['trx_old'] == 0)]
#xpopd_w_trx_zero_sales

xpopd_w_trx.to_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\temp\xpopd_w_trx.xlsx" %(prid),sheet_name = "xpopd_w_trx")

if xpopd_w_trx_deviations.empty == True:
 xpopd_w_trx_deviations_flag = 'Y'
else:
 xpopd_w_trx_deviations.to_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\temp\deviations_trx.xlsx" %(prid),sheet_name = "deviations_trx")
 
if xpopd_w_trx_zero_sales.empty == True:
 xpopd_w_trx_zero_sales_flag = 'Y'
else:
 xpopd_w_trx_zero_sales.to_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\temp\zero_sales_trx.xlsx" %(prid),sheet_name = "zero_sales_trx")

##Zero_sales_trx:

if xpopd_w_trx_zero_sales_flag == 'Y':
  pass
else:
 Zero_sales_trx_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\temp\zero_sales_trx.xlsx" %(prid))
 xpopd_w_zero_sales_trx = Zero_sales_trx_df.fillna(method = "ffill")
#xpopd_w_zero_sales_trx

 xpopd_w_zero_sales_trx.loc[(xpopd_w_zero_sales_trx['trx_old'] == 0) & (xpopd_w_zero_sales_trx['trx_new'] == 0), 'comments'] = "no comments"
 xpopd_w_zero_sales_trx.loc[(xpopd_w_zero_sales_trx['trx_new'] == 0) & (xpopd_w_zero_sales_trx['trx_old'] != 0), 'comments'] = xpopd_w_zero_sales_trx['prod_nm'] + " sales dropped to zero under " + xpopd_w_zero_sales_trx["mkt_nm"]
 xpopd_w_zero_sales_trx.loc[(xpopd_w_zero_sales_trx['trx_old'] == 0) & (xpopd_w_zero_sales_trx['trx_new'] != 0), 'comments'] = xpopd_w_zero_sales_trx['prod_nm'] + " is flowing under " + xpopd_w_zero_sales_trx["mkt_nm"]
 
 xpopd_w_zero_sales_trx = xpopd_w_zero_sales_trx.sort_values(by='prod_nm')

##Deviation_trx:

if xpopd_w_trx_deviations_flag == 'Y':
  pass
else:
 Deviation_trx_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\temp\deviations_trx.xlsx" %(prid))
 xpopd_w_deviation_trx = Deviation_trx_df.fillna(method = "ffill")
#xpopd_w_deviation_trx

#deviation_trx['deviation']= deviation_trx['deviation'].astype(str)
 
 xpopd_w_deviation_trx.loc[xpopd_w_deviation_trx['deviation'] > 5, 'comments'] = xpopd_w_deviation_trx['brd_nm'] + " sales is increased by " + xpopd_w_deviation_trx['deviation'].map(str) + "%" + " in " + xpopd_w_deviation_trx['mkt_nm']
 xpopd_w_deviation_trx.loc[xpopd_w_deviation_trx['deviation'] < -5, 'comments'] = xpopd_w_deviation_trx['brd_nm'] + " sales is dropped by " + xpopd_w_deviation_trx['deviation'].map(str) + "%" + " in " + xpopd_w_deviation_trx['mkt_nm']
 
 xpopd_w_deviation_trx = xpopd_w_deviation_trx.sort_values(by='prod_nm')
 
 ####################################Wrting the output to excel file################################################################### 	

if [xpopd_w_nrx_deviations_flag == 'Y'] or [xpopd_w_nrx_zero_sales_flag == 'Y'] or [xpopd_w_trx_deviations_flag == 'Y'] or [xpopd_w_trx_zero_sales_flag == 'Y']:
    with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\output\xpopd_w_output.xlsx" %(prid)) as writer:    
      xpopd_w_nrx.to_excel(writer,sheet_name = "xpopd_w_nrx")
      xpopd_w_trx.to_excel(writer,sheet_name = "xpopd_w_trx")
    
    if xpopd_w_nrx_deviations_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\output\xpopd_w_output.xlsx" %(prid))
        book.create_sheet("deviation_nrx")
        sheet = book["deviation_nrx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\output\xpopd_w_output.xlsx" %(prid))
        
    else:
        with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\output\xpopd_w_output.xlsx" %(prid), engine='openpyxl', mode='a') as writer: 
         xpopd_w_deviation_nrx.to_excel(writer,sheet_name = "deviation_nrx")
        
    if xpopd_w_nrx_zero_sales_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\output\xpopd_w_output.xlsx" %(prid))
        book.create_sheet("Zero_sales_nrx")
        sheet = book["Zero_sales_nrx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\output\xpopd_w_output.xlsx" %(prid))
        
    else:
        with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\output\xpopd_w_output.xlsx" %(prid), engine='openpyxl', mode='a') as writer: 
         xpopd_w_zero_sales_nrx.to_excel(writer,sheet_name = "zero_sales_nrx")
    
    if xpopd_w_trx_deviations_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\output\xpopd_w_output.xlsx" %(prid))
        book.create_sheet("deviation_trx")
        sheet = book["deviation_trx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\output\xpopd_w_output.xlsx" %(prid))
        
    else:
        with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\output\xpopd_w_output.xlsx" %(prid), engine='openpyxl', mode='a') as writer: 
         xpopd_w_deviation_trx.to_excel(writer,sheet_name = "deviation_trx")
  

    if xpopd_w_trx_zero_sales_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\output\xpopd_w_output.xlsx" %(prid))
        book.create_sheet("Zero_sales_trx")
        sheet = book["Zero_sales_trx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\output\xpopd_w_output.xlsx" %(prid))
        
    else:
        with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\output\xpopd_w_output.xlsx" %(prid), engine='openpyxl', mode='a') as writer: 
         xpopd_w_zero_sales_trx.to_excel(writer,sheet_name = "Zero_sales_trx")

##################################################XPOPD_WEEKLY_Data_Visualization#########################################################################

####################################################MARKET##################################################################################

if xpopd_w_nrx_deviations_flag == 'Y':
  pass
else:
 xpopd_w_mkt_dev = xpopd_w_deviation_nrx[['mkt_nm', 'deviation']].copy()
 #xpopd_w_mkt_dev

 xpopd_w_mkt_dev.to_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\temp\xpopd_w_mkt_deviation.xlsx" %(prid),index=False)
  
 xpopd_w_market_deviation = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\temp\xpopd_w_mkt_deviation.xlsx" %(prid))
 xpopd_w_market_deviation

 xpopd_w_mar_res = xpopd_w_market_deviation.groupby('mkt_nm')
 #xpopd_w_mar_res

 xpopd_w_mar_out = xpopd_w_mar_res.mean()
 #xpopd_w_mar_out

 xpopd_w_mar_out.to_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\temp\xpopd_w_mkt_dev_output.xlsx" %(prid))

 xpopd_w_mar_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\temp\xpopd_w_mkt_dev_output.xlsx" %(prid))
  
 sns.set(rc={'figure.figsize':(11.7,8.27)})
 xpopd_w_mar_sns_plot = sns.barplot(x="deviation",y="mkt_nm",data=xpopd_w_mar_df)
 xpopd_w_mar_sns_plot.figure.savefig(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\temp\xpopd_w_mar_output.pdf" %(prid),bbox_inches="tight",pad_inches=1)

#######################################BRAND######################################################################################

 xpopd_w_brand_dev = xpopd_w_deviation_nrx[['brd_nm', 'deviation']].copy()
 #xpopd_w_brand_dev

 xpopd_w_brand_dev.to_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\temp\xpopd_w_brand_deviation.xlsx" %(prid),index=False)

 xpopd_w_brand_deviation = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\temp\xpopd_w_brand_deviation.xlsx" %(prid))
 xpopd_w_brand_deviation

 xpopd_w_brand_res = xpopd_w_brand_deviation.groupby('brd_nm')
 #xpopd_w_brand_res

 xpopd_w_brand_out = xpopd_w_brand_res.mean()
 #xpopd_w_brand_out

 xpopd_w_brand_out.to_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\temp\xpopd_w_brand_dev_output.xlsx" %(prid))

 xpopd_w_brand_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\temp\xpopd_w_brand_dev_output.xlsx" %(prid))

 sns.set(rc={'figure.figsize':(11.7,8.27)})
 xpopd_w_brand_sns_plot = sns.barplot(x="deviation",y="brd_nm",data=xpopd_w_brand_df)
 xpopd_w_brand_sns_plot.figure.savefig(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\temp\xpopd_w_brand_output.pdf" %(prid),bbox_inches="tight",pad_inches=1)

###############################################PRODUCT##########################################################################

 xpopd_w_product_dev = xpopd_w_deviation_nrx[['prod_nm', 'deviation']].copy()
 #xpopd_w_product_dev

 xpopd_w_product_dev.to_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\temp\xpopd_w_product_deviation.xlsx" %(prid),index=False)

 xpopd_w_product_deviation = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\temp\xpopd_w_product_deviation.xlsx" %(prid))
 xpopd_w_product_deviation

 xpopd_w_product_res = xpopd_w_product_deviation.groupby('prod_nm')
 #xpopd_w_product_res

 xpopd_w_product_out = xpopd_w_product_res.mean()
 #xpopd_w_product_out

 xpopd_w_product_out.to_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\temp\xpopd_w_product_dev_output.xlsx" %(prid))

 xpopd_w_product_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\temp\xpopd_w_product_dev_output.xlsx" %(prid))

 sns.set(rc={'figure.figsize':(11.7,8.27)})
 xpopd_w_prod_sns_plot = sns.barplot(x="deviation",y="prod_nm",data=xpopd_w_product_df)
 xpopd_w_prod_sns_plot.figure.savefig(r"C:\Users\%s\Desktop\unaligned\xpopd_weekly\temp\xpopd_w_product_output.pdf" %(prid),bbox_inches="tight",pad_inches=1)

#################################Moving the image files to output folder######################################################

if xpopd_w_nrx_deviations_flag == 'Y':
  pass
else:
 sourcepath = r'C:\Users\%s\Desktop\unaligned\xpopd_weekly\temp\\' %(prid)
 sourcefiles = os.listdir(sourcepath)
 destinationpath = r'C:\Users\%s\Desktop\unaligned\xpopd_weekly\output\\' %(prid)
 for file in sourcefiles:
    if file.endswith('.pdf'):
        shutil.move(os.path.join(sourcepath,file), os.path.join(destinationpath,file))

###################################################Deleting unwanted files after run from temp folder########################################################

import os
import shutil

for root, dirs, files in os.walk(r'C:\Users\%s\Desktop\unaligned\xpopd_weekly\temp\\' %(prid)):
    for f in files:
        os.unlink(os.path.join(root, f))
    for d in dirs:
        shutil.rmtree(os.path.join(root, d))
        
#######################################################XPOPD_MONTHLY##########################################################


#########################Nrx##################################################################################################################

xpopd_m_nrx = input_xpopd_m_df.pivot_table(index = ["mkt_nm","brd_nm","prod_nm"],values = ["nrx_new","nrx_old"])
xpopd_m_nrx["deviation"] = ((xpopd_m_nrx.nrx_new - xpopd_m_nrx.nrx_old)/xpopd_m_nrx.nrx_old)*100
#xpopd_m_nrx
xpopd_m_nrx["deviation"] = xpopd_m_nrx["deviation"].round(round(1))
#xpopd_m_nrx
xpopd_m_nrx_deviation = xpopd_m_nrx[((xpopd_m_nrx['deviation'] <= -5) | (xpopd_m_nrx['deviation'] >=5))]
#xpopd_m_nrx_deviation
xpopd_m_nrx_deviations = xpopd_m_nrx_deviation[((xpopd_m_nrx_deviation['nrx_new'] != 0) & (xpopd_m_nrx_deviation['nrx_old'] != 0))]
#xpopd_m_nrx_deviations
xpopd_m_nrx_zero_sales = xpopd_m_nrx[(xpopd_m_nrx['nrx_new'] == 0) | (xpopd_m_nrx['nrx_old'] == 0)]
#xpopd_m_nrx_zero_sales

xpopd_m_nrx.to_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\temp\xpopd_m_nrx.xlsx" %(prid),sheet_name = "xpopd_m_nrx")

xpopd_m_nrx_deviations_flag = 'N'
xpopd_m_nrx_zero_sales_flag = 'N'
xpopd_m_trx_deviations_flag = 'N'
xpopd_m_trx_zero_sales_flag = 'N'

if xpopd_m_nrx_deviations.empty == True:
 xpopd_m_nrx_deviations_flag = 'Y'
else:
 xpopd_m_nrx_deviations.to_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\temp\deviations_nrx.xlsx" %(prid),sheet_name = "deviations_nrx")

if xpopd_m_nrx_zero_sales.empty == True:
 xpopd_m_nrx_zero_sales_flag = 'Y'
else:
 xpopd_m_nrx_zero_sales.to_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\temp\zero_sales_nrx.xlsx" %(prid),sheet_name = "zero_sales_nrx")

##Zero_sales_nrx:

if xpopd_m_nrx_zero_sales_flag == 'Y':
  pass
else:
 zs_xpopd_m_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\temp\zero_sales_nrx.xlsx" %(prid))
 xpopd_m_zero_sales_nrx = zs_xpopd_m_df.fillna(method = "ffill")
#xpopd_m_zero_sales_nrx

 xpopd_m_zero_sales_nrx.loc[(xpopd_m_zero_sales_nrx['nrx_old'] == 0) & (xpopd_m_zero_sales_nrx['nrx_new'] == 0), 'comments'] = "no comments"
 xpopd_m_zero_sales_nrx.loc[(xpopd_m_zero_sales_nrx['nrx_new'] == 0) & (xpopd_m_zero_sales_nrx['nrx_old'] != 0), 'comments'] = xpopd_m_zero_sales_nrx['prod_nm'] + " sales dropped to zero under " + xpopd_m_zero_sales_nrx["mkt_nm"]
 xpopd_m_zero_sales_nrx.loc[(xpopd_m_zero_sales_nrx['nrx_old'] == 0) & (xpopd_m_zero_sales_nrx['nrx_new'] != 0), 'comments'] = xpopd_m_zero_sales_nrx['prod_nm'] + " is flowing under " + xpopd_m_zero_sales_nrx["mkt_nm"]
 
 xpopd_m_zero_sales_nrx = xpopd_m_zero_sales_nrx.sort_values(by='prod_nm')

##Deviation_nrx:

if xpopd_m_nrx_deviations_flag == 'Y':
  pass
else:
 dev_xpopd_m_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\temp\deviations_nrx.xlsx" %(prid))
 xpopd_m_deviation_nrx = dev_xpopd_m_df.fillna(method = "ffill")
#xpopd_m_deviation_nrx

#xpopd_m_deviation_nrx['deviation']= deviation_nrx['deviation'].astype(str)

 xpopd_m_deviation_nrx.loc[xpopd_m_deviation_nrx['deviation'] > 5, 'comments'] = xpopd_m_deviation_nrx['brd_nm'] + " sales is increased by " + xpopd_m_deviation_nrx['deviation'].map(str) + "%" + " in " + xpopd_m_deviation_nrx['mkt_nm']
 xpopd_m_deviation_nrx.loc[xpopd_m_deviation_nrx['deviation'] < -5, 'comments'] = xpopd_m_deviation_nrx['brd_nm'] + " sales is dropped by " + xpopd_m_deviation_nrx['deviation'].map(str) + "%" + " in " + xpopd_m_deviation_nrx['mkt_nm']
 
 xpopd_m_deviation_nrx = xpopd_m_deviation_nrx.sort_values(by='prod_nm')

#############################################Trx##############################################################################################

xpopd_m_trx = input_xpopd_m_df.pivot_table(index = ["mkt_nm","brd_nm","prod_nm"],values = ["trx_new","trx_old"])
xpopd_m_trx["deviation"] = ((xpopd_m_trx.trx_new - xpopd_m_trx.trx_old)/xpopd_m_trx.trx_old)*100
#xpopd_m_trx
xpopd_m_trx["deviation"] = xpopd_m_trx["deviation"].round(round(1))
#xpopd_m_trx
xpopd_m_trx_deviation = xpopd_m_trx[((xpopd_m_trx['deviation'] <= -5) | (xpopd_m_trx['deviation'] >=5))]
#xpopd_m_trx_deviation
xpopd_m_trx_deviations = xpopd_m_trx_deviation[((xpopd_m_trx_deviation['trx_new'] != 0) & (xpopd_m_trx_deviation['trx_old'] != 0))]
#xpopd_m_trx_deviations
xpopd_m_trx_zero_sales = xpopd_m_trx[(xpopd_m_trx['trx_new'] == 0) | (xpopd_m_trx['trx_old'] == 0)]
#xpopd_m_trx_zero_sales

xpopd_m_trx.to_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\temp\xpopd_m_trx.xlsx" %(prid),sheet_name = "xpopd_m_trx")

if xpopd_m_trx_deviations.empty == True:
 xpopd_m_trx_deviations_flag = 'Y'
else:
 xpopd_m_trx_deviations.to_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\temp\deviations_trx.xlsx" %(prid),sheet_name = "deviations_trx")
 
if xpopd_m_trx_zero_sales.empty == True:
 xpopd_m_trx_zero_sales_flag = 'Y'
else:
 xpopd_m_trx_zero_sales.to_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\temp\zero_sales_trx.xlsx" %(prid),sheet_name = "zero_sales_trx")

##Zero_sales_trx:

if xpopd_m_trx_zero_sales_flag == 'Y':
  pass
else:
 Zero_sales_trx_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\temp\zero_sales_trx.xlsx" %(prid))
 xpopd_m_zero_sales_trx = Zero_sales_trx_df.fillna(method = "ffill")
#xpopd_m_zero_sales_trx

 xpopd_m_zero_sales_trx.loc[(xpopd_m_zero_sales_trx['trx_old'] == 0) & (xpopd_m_zero_sales_trx['trx_new'] == 0), 'comments'] = "no comments"
 xpopd_m_zero_sales_trx.loc[(xpopd_m_zero_sales_trx['trx_new'] == 0) & (xpopd_m_zero_sales_trx['trx_old'] != 0), 'comments'] = xpopd_m_zero_sales_trx['prod_nm'] + " sales dropped to zero under " + xpopd_m_zero_sales_trx["mkt_nm"]
 xpopd_m_zero_sales_trx.loc[(xpopd_m_zero_sales_trx['trx_old'] == 0) & (xpopd_m_zero_sales_trx['trx_new'] != 0), 'comments'] = xpopd_m_zero_sales_trx['prod_nm'] + " is flowing under " + xpopd_m_zero_sales_trx["mkt_nm"]
 
 xpopd_m_zero_sales_trx = xpopd_m_zero_sales_trx.sort_values(by='prod_nm')

##Deviation_trx:

if xpopd_m_trx_deviations_flag == 'Y':
  pass
else:
 Deviation_trx_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\temp\deviations_trx.xlsx" %(prid))
 xpopd_m_deviation_trx = Deviation_trx_df.fillna(method = "ffill")
#xpopd_m_deviation_trx

#deviation_trx['deviation']= deviation_trx['deviation'].astype(str)
 
 xpopd_m_deviation_trx.loc[xpopd_m_deviation_trx['deviation'] > 5, 'comments'] = xpopd_m_deviation_trx['brd_nm'] + " sales is increased by " + xpopd_m_deviation_trx['deviation'].map(str) + "%" + " in " + xpopd_m_deviation_trx['mkt_nm']
 xpopd_m_deviation_trx.loc[xpopd_m_deviation_trx['deviation'] < -5, 'comments'] = xpopd_m_deviation_trx['brd_nm'] + " sales is dropped by " + xpopd_m_deviation_trx['deviation'].map(str) + "%" + " in " + xpopd_m_deviation_trx['mkt_nm']
 
 xpopd_m_deviation_trx = xpopd_m_deviation_trx.sort_values(by='prod_nm')
 
 ####################################Wrting the output to excel file################################################################### 	

if [xpopd_m_nrx_deviations_flag == 'Y'] or [xpopd_m_nrx_zero_sales_flag == 'Y'] or [xpopd_m_trx_deviations_flag == 'Y'] or [xpopd_m_trx_zero_sales_flag == 'Y']:
    with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\output\xpopd_m_output.xlsx" %(prid)) as writer:    
      xpopd_m_nrx.to_excel(writer,sheet_name = "xpopd_m_nrx")
      xpopd_m_trx.to_excel(writer,sheet_name = "xpopd_m_trx")
    
    if xpopd_m_nrx_deviations_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\output\xpopd_m_output.xlsx" %(prid))
        book.create_sheet("deviation_nrx")
        sheet = book["deviation_nrx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\output\xpopd_m_output.xlsx" %(prid))
        
    else:
        with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\output\xpopd_m_output.xlsx" %(prid), engine='openpyxl', mode='a') as writer: 
         xpopd_m_deviation_nrx.to_excel(writer,sheet_name = "deviation_nrx")   

    if xpopd_m_nrx_zero_sales_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\output\xpopd_m_output.xlsx" %(prid))
        book.create_sheet("Zero_sales_nrx")
        sheet = book["Zero_sales_nrx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\output\xpopd_m_output.xlsx" %(prid))
        
    else:
        with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\output\xpopd_m_output.xlsx" %(prid), engine='openpyxl', mode='a') as writer: 
         xpopd_m_zero_sales_nrx.to_excel(writer,sheet_name = "zero_sales_nrx")
    
    if xpopd_m_trx_deviations_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\output\xpopd_m_output.xlsx" %(prid))
        book.create_sheet("deviation_trx")
        sheet = book["deviation_trx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\output\xpopd_m_output.xlsx" %(prid))
        
    else:
        with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\output\xpopd_m_output.xlsx" %(prid), engine='openpyxl', mode='a') as writer: 
         xpopd_m_deviation_trx.to_excel(writer,sheet_name = "deviation_trx")
  

    if xpopd_m_trx_zero_sales_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\output\xpopd_m_output.xlsx" %(prid))
        book.create_sheet("zero_sales_trx")
        sheet = book["zero_sales_trx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\output\xpopd_m_output.xlsx" %(prid))
        
    else:
        with pd.ExcelWriter(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\output\xpopd_m_output.xlsx" %(prid), engine='openpyxl', mode='a') as writer: 
         xpopd_m_zero_sales_trx.to_excel(writer,sheet_name = "zero_sales_trx")


##################################################XPOPD_MONTHLY_Data_Visualization#########################################################################

####################################################MARKET##################################################################################


if xpopd_m_nrx_deviations_flag == 'Y':
  pass
else:
 xpopd_m_mkt_dev = xpopd_m_deviation_nrx[['mkt_nm', 'deviation']].copy()
 #xpopd_m_mkt_dev

 xpopd_m_mkt_dev.to_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\temp\xpopd_m_mkt_deviation.xlsx" %(prid),index=False)
 
 xpopd_m_market_deviation = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\temp\xpopd_m_mkt_deviation.xlsx" %(prid))
 xpopd_m_market_deviation

 xpopd_m_mar_res = xpopd_m_market_deviation.groupby('mkt_nm')
 #xpopd_m_mar_res

 xpopd_m_mar_out = xpopd_m_mar_res.mean()
 #xpopd_m_mar_out

 xpopd_m_mar_out.to_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\temp\xpopd_m_mkt_dev_output.xlsx" %(prid))

 xpopd_m_mar_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\temp\xpopd_m_mkt_dev_output.xlsx" %(prid))

 sns.set(rc={'figure.figsize':(11.7,8.27)})
 xpopd_m_mar_sns_plot = sns.barplot(x="deviation",y="mkt_nm",data=xpopd_m_mar_df)
 xpopd_m_mar_sns_plot.figure.savefig(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\temp\xpopd_m_mar_output.pdf" %(prid),bbox_inches="tight",pad_inches=1)

#######################################BRAND######################################################################################

 xpopd_m_brand_dev = xpopd_m_deviation_nrx[['brd_nm', 'deviation']].copy()
 #xpopd_m_brand_dev

 xpopd_m_brand_dev.to_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\temp\xpopd_m_brand_deviation.xlsx" %(prid),index=False)

 xpopd_m_brand_deviation = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\temp\xpopd_m_brand_deviation.xlsx" %(prid))
 xpopd_m_brand_deviation

 xpopd_m_brand_res = xpopd_m_brand_deviation.groupby('brd_nm')
 #xpopd_m_brand_res

 xpopd_m_brand_out = xpopd_m_brand_res.mean()
 #xpopd_m_brand_out

 xpopd_m_brand_out.to_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\temp\xpopd_m_brand_dev_output.xlsx" %(prid))

 xpopd_m_brand_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\temp\xpopd_m_brand_dev_output.xlsx" %(prid))

 sns.set(rc={'figure.figsize':(11.7,8.27)})
 xpopd_m_brand_sns_plot = sns.barplot(x="deviation",y="brd_nm",data=xpopd_m_brand_df)
 xpopd_m_brand_sns_plot.figure.savefig(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\temp\xpopd_m_brand_output.pdf" %(prid),bbox_inches="tight",pad_inches=1)

###############################################PRODUCT##########################################################################

 xpopd_m_product_dev = xpopd_m_deviation_nrx[['prod_nm', 'deviation']].copy()
 #xpopd_m_product_dev

 xpopd_m_product_dev.to_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\temp\xpopd_m_product_deviation.xlsx" %(prid),index=False)

 xpopd_m_product_deviation = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\temp\xpopd_m_product_deviation.xlsx" %(prid))
 xpopd_m_product_deviation

 xpopd_m_product_res = xpopd_m_product_deviation.groupby('prod_nm')
 #xpopd_m_product_res

 xpopd_m_product_out = xpopd_m_product_res.mean()
 #xpopd_m_product_out

 xpopd_m_product_out.to_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\temp\xpopd_m_product_dev_output.xlsx" %(prid))

 xpopd_m_product_df = pd.read_excel(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\temp\xpopd_m_product_dev_output.xlsx" %(prid))

 sns.set(rc={'figure.figsize':(11.7,8.27)})
 xpopd_m_prod_sns_plot = sns.barplot(x="deviation",y="prod_nm",data=xpopd_m_product_df)
 xpopd_m_prod_sns_plot.figure.savefig(r"C:\Users\%s\Desktop\unaligned\xpopd_monthly\temp\xpopd_m_product_output.pdf" %(prid),bbox_inches="tight",pad_inches=1)

#################################Moving the image files to output folder######################################################

if xpopd_m_nrx_deviations_flag == 'Y':
  pass
else:
 sourcepath = r'C:\Users\%s\Desktop\unaligned\xpopd_monthly\temp\\' %(prid)
 sourcefiles = os.listdir(sourcepath)
 destinationpath = r'C:\Users\%s\Desktop\unaligned\xpopd_monthly\output\\' %(prid)
 for file in sourcefiles:
    if file.endswith('.pdf'):
        shutil.move(os.path.join(sourcepath,file), os.path.join(destinationpath,file))

###################################################Deleting unwanted files after run from temp folder########################################################

import os
import shutil

for root, dirs, files in os.walk(r'C:\Users\%s\Desktop\unaligned\xpopd_monthly\temp\\' %(prid)):
    for f in files:
        os.unlink(os.path.join(root, f))
    for d in dirs:
        shutil.rmtree(os.path.join(root, d))


# In[ ]:


# In[ ]:




