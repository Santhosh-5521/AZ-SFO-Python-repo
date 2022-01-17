#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
import os 
from os import listdir

######################Reading input files from local and storing it in dataframes##############################################################

input_ddd_w_df = pd.read_excel(r"C:\Users\ksxp996\Desktop\unaligned\ddd_weekly\input\f_sls_outl_prod_wk.xlsx")
input_ddd_m_df = pd.read_excel(r"C:\Users\ksxp996\Desktop\unaligned\ddd_monthly\input\f_sls_outl_prod_mnth.xlsx")
input_xpo_w_df = pd.read_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpo_weekly\input\f_sls_hcp_prod_plan_wk.xlsx")
input_xpopd_w_df = pd.read_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_weekly\input\f_sls_hcp_prod_plan_dyn_wk.xlsx")
input_xpopd_m_df = pd.read_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_monthly\input\f_sls_hcp_prod_plan_dyn_mnth.xlsx")

###################################################DDD WEEKLY############################################################################################

##########################UNITS#############################################################################

ddd_w_units = input_ddd_w_df.pivot_table(index = ["mkt_nm","brd_nm","prod_nm"],values = ["tot_units_new","tot_units_old"])
ddd_w_units["deviation"] = ((ddd_w_units.tot_units_new - ddd_w_units.tot_units_old)/ddd_w_units.tot_units_old)*100
#ddd_w_units
ddd_w_units_deviation = ddd_w_units[((ddd_w_units['deviation'] <= -5) | (ddd_w_units['deviation'] >=5))]
#ddd_w_units_deviation
ddd_w_units_deviations = ddd_w_units_deviation[((ddd_w_units_deviation['tot_units_new'] != 0) & (ddd_w_units_deviation['tot_units_old'] != 0))]
#ddd_w_units_deviations
ddd_w_units_zero_sales = ddd_w_units[(ddd_w_units['tot_units_new'] == 0) | (ddd_w_units['tot_units_old'] == 0)]
#ddd_w_units_zero_sales

ddd_w_units.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\ddd_weekly\ddd_w_units.xlsx",sheet_name = "ddd_w_units")

ddd_w_units_deviations_flag = 'N'
ddd_w_units_zero_sales_flag = 'N'
ddd_w_dollars_deviations_flag = 'N'
ddd_w_dollars_zero_sales_flag = 'N'

if ddd_w_units_deviations.empty == True:
 ddd_w_units_deviations_flag = 'Y'
else:
 ddd_w_units_deviations.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\ddd_weekly\deviations_units.xlsx",sheet_name = "deviations_units")

if ddd_w_units_zero_sales.empty == True:
 ddd_w_units_zero_sales_flag = 'Y'
else:
 ddd_w_units_zero_sales.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\ddd_weekly\zero_sales_units.xlsx",sheet_name = "zero_sales_units")

##Zero_sales_units:

if ddd_w_units_deviations_flag == 'Y':
  pass
else:
 zs_ddd_w_df = pd.read_excel(r"C:\Users\ksxp996\Desktop\unaligned\ddd_weekly\zero_sales_units.xlsx")
 ddd_w_zero_sales_units = zs_ddd_w_df.fillna(method = "ffill")
#zero_sales_units

 ddd_w_zero_sales_units.loc[(ddd_w_zero_sales_units['tot_units_old'] == 0) & (ddd_w_zero_sales_units['tot_units_new'] == 0), 'comments'] = "no comments"
 ddd_w_zero_sales_units.loc[(ddd_w_zero_sales_units['tot_units_new'] == 0) & (ddd_w_zero_sales_units['tot_units_old'] != 0), 'comments'] = ddd_w_zero_sales_units['prod_nm'] + " sales dropped to zero under " + ddd_w_zero_sales_units["mkt_nm"]
 ddd_w_zero_sales_units.loc[(ddd_w_zero_sales_units['tot_units_old'] == 0) & (ddd_w_zero_sales_units['tot_units_new'] != 0), 'comments'] = ddd_w_zero_sales_units['prod_nm'] + " is flowing under " + ddd_w_zero_sales_units["mkt_nm"]

##Deviation_units:

if ddd_w_units_zero_sales_flag == 'Y':
  pass
else:
 dev_ddd_w_df = pd.read_excel(r"C:\Users\ksxp996\Desktop\unaligned\ddd_weekly\deviations_units.xlsx")
 ddd_w_deviation_units = dev_ddd_w_df.fillna(method = "ffill")
#ddd_w_deviation_units

#deviation_units['deviation']= deviation_units['deviation'].astype(str)

 ddd_w_deviation_units.loc[ddd_w_deviation_units['deviation'] > 5, 'comments'] = ddd_w_deviation_units['brd_nm'] + " sales is increased in " + ddd_w_deviation_units["mkt_nm"]
 ddd_w_deviation_units.loc[ddd_w_deviation_units['deviation'] < -5, 'comments'] = ddd_w_deviation_units['brd_nm'] + " sales is dropped in " + ddd_w_deviation_units["mkt_nm"]

#############################################DOLLARS##############################################################################################

ddd_w_dollars = input_ddd_w_df.pivot_table(index = ["mkt_nm","brd_nm","prod_nm"],values = ["tot_dolrs_new","tot_dolrs_old"])
ddd_w_dollars["deviation"] = ((ddd_w_dollars.tot_dolrs_new - ddd_w_dollars.tot_dolrs_old)/ddd_w_dollars.tot_dolrs_old)*100
#ddd_w_dollars
ddd_w_dollars_deviation = ddd_w_dollars[((ddd_w_dollars['deviation'] <= -5) | (ddd_w_dollars['deviation'] >=5))]
#ddd_w_dollars_deviation
ddd_w_dollars_deviations = ddd_w_dollars_deviation[((ddd_w_dollars_deviation['tot_dolrs_new'] != 0) & (ddd_w_dollars_deviation['tot_dolrs_old'] != 0))]
#ddd_w_dollars_deviations
ddd_w_dollars_zero_sales = ddd_w_dollars[(ddd_w_dollars['tot_dolrs_new'] == 0) | (ddd_w_dollars['tot_dolrs_old'] == 0)]
#ddd_w_dollars_zero_sales

ddd_w_dollars.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\ddd_weekly\ddd_w_dollars.xlsx",sheet_name = "ddd_w_dollars")

if ddd_w_dollars_deviations.empty == True:
 ddd_w_dollars_deviations_flag = 'Y'
else:
 ddd_w_dollars_deviations.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\ddd_weekly\deviations_dollars.xlsx",sheet_name = "deviations_dollars")
 
if ddd_w_dollars_zero_sales.empty == True:
 ddd_w_dollars_zero_sales_flag = 'Y'
else:
 ddd_w_dollars_zero_sales.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\ddd_weekly\zero_sales_dollars.xlsx",sheet_name = "zero_sales_dollars")

##Zero_sales_dollars:

if ddd_w_dollars_deviations_flag == 'Y':
  pass
else:
 Zero_sales_dollars_df = pd.read_excel(r"C:\Users\ksxp996\Desktop\unaligned\ddd_weekly\zero_sales_dollars.xlsx")
 ddd_w_zero_sales_dollars = Zero_sales_dollars_df.fillna(method = "ffill")
#ddd_w_zero_sales_dollars

 ddd_w_zero_sales_dollars.loc[(ddd_w_zero_sales_dollars['tot_dolrs_old'] == 0) & (ddd_w_zero_sales_dollars['tot_dolrs_new'] == 0), 'comments'] = "no comments"
 ddd_w_zero_sales_dollars.loc[(ddd_w_zero_sales_dollars['tot_dolrs_new'] == 0) & (ddd_w_zero_sales_dollars['tot_dolrs_old'] != 0), 'comments'] = ddd_w_zero_sales_dollars['prod_nm'] + " sales dropped to zero under " + ddd_w_zero_sales_dollars["mkt_nm"]
 ddd_w_zero_sales_dollars.loc[(ddd_w_zero_sales_dollars['tot_dolrs_old'] == 0) & (ddd_w_zero_sales_dollars['tot_dolrs_new'] != 0), 'comments'] = ddd_w_zero_sales_dollars['prod_nm'] + " is flowing under " + ddd_w_zero_sales_dollars["mkt_nm"]

##Deviation_dollars:

if ddd_w_dollars_zero_sales_flag == 'Y':
  pass
else:
 Deviation_dollars_df = pd.read_excel(r"C:\Users\ksxp996\Desktop\unaligned\ddd_weekly\deviations_dollars.xlsx")
 ddd_w_deviation_dollars = Deviation_dollars_df.fillna(method = "ffill")
#deviation_dollars

#deviation_dollars['deviation']= deviation_dollars['deviation'].astype(str)

 ddd_w_deviation_dollars.loc[ddd_w_deviation_dollars['deviation'] > 5, 'comments'] = ddd_w_deviation_dollars['brd_nm'] + " sales is increased in " + ddd_w_deviation_dollars["mkt_nm"]
 ddd_w_deviation_dollars.loc[ddd_w_deviation_dollars['deviation'] < -5, 'comments'] = ddd_w_deviation_dollars['brd_nm'] + " sales is dropped in " + ddd_w_deviation_dollars["mkt_nm"]
 
#########################Writing the output to excel file#########################################################################

if [ddd_w_units_deviations_flag == 'Y'] or [ddd_w_units_zero_sales_flag == 'Y'] or [ddd_w_dollars_deviations_flag == 'Y'] or [ddd_w_dollars_zero_sales_flag == 'Y']:
    with pd.ExcelWriter(r"C:\Users\ksxp996\Desktop\unaligned\ddd_weekly\output\ddd_w_output.xlsx") as writer:    
      ddd_w_units.to_excel(writer,sheet_name = "ddd_w_units")
      ddd_w_dollars.to_excel(writer,sheet_name = "ddd_w_dollars")
    
    if ddd_w_units_deviations_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\ksxp996\Desktop\unaligned\ddd_weekly\output\ddd_w_output.xlsx")
        book.create_sheet("deviation_units")
        sheet = book["deviation_units"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\ksxp996\Desktop\unaligned\ddd_weekly\output\ddd_w_output.xlsx")
        
    else:
        with pd.ExcelWriter(r"C:\Users\ksxp996\Desktop\unaligned\ddd_weekly\output\ddd_w_output.xlsx", engine='openpyxl', mode='a') as writer: 
         ddd_w_deviation_units.to_excel(writer,sheet_name = "deviation_units")

    if ddd_w_units_zero_sales_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\ksxp996\Desktop\unaligned\ddd_weekly\output\ddd_w_output.xlsx")
        book.create_sheet("Zero_sales_units")
        sheet = book["Zero_sales_units"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\ksxp996\Desktop\unaligned\ddd_weekly\output\ddd_w_output.xlsx")
        
    else:
        with pd.ExcelWriter(r"C:\Users\ksxp996\Desktop\unaligned\ddd_weekly\output\ddd_w_output.xlsx", engine='openpyxl', mode='a') as writer: 
         ddd_w_zero_sales_units.to_excel(writer,sheet_name = "Zero_sales_units")  
    
    if ddd_w_dollars_deviations_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\ksxp996\Desktop\unaligned\ddd_weekly\output\ddd_w_output.xlsx")
        book.create_sheet("deviation_dollars")
        sheet = book["deviation_dollars"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\ksxp996\Desktop\unaligned\ddd_weekly\output\ddd_w_output.xlsx")
        
    else:
        with pd.ExcelWriter(r"C:\Users\ksxp996\Desktop\unaligned\ddd_weekly\output\ddd_w_output.xlsx", engine='openpyxl', mode='a') as writer: 
         ddd_w_deviation_dollars.to_excel(writer,sheet_name = "deviation_dollars")
  

    if ddd_w_dollars_zero_sales_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\ksxp996\Desktop\unaligned\ddd_weekly\output\ddd_w_output.xlsx")
        book.create_sheet("Zero_sales_dollars")
        sheet = book["Zero_sales_dollars"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\ksxp996\Desktop\unaligned\ddd_weekly\output\ddd_w_output.xlsx")
        
    else:
        with pd.ExcelWriter(r"C:\Users\ksxp996\Desktop\unaligned\ddd_weekly\output\ddd_w_output.xlsx", engine='openpyxl', mode='a') as writer: 
         ddd_w_zero_sales_dollars.to_excel(writer,sheet_name = "Zero_sales_dollars")

ddd_w_folder_path = r'C:\Users\ksxp996\Desktop\unaligned\ddd_weekly\\'

for file_name in listdir(ddd_w_folder_path):
    
    if file_name.endswith('.xlsx'):
        
        os.remove(ddd_w_folder_path + file_name)
        
##################################################DDD_MONTHLY#######################################################################


#########################UNITS##################################################################################################################

ddd_m_units = input_ddd_m_df.pivot_table(index = ["mkt_nm","brd_nm","prod_nm"],values = ["tot_units_new","tot_units_old"])
ddd_m_units["deviation"] = ((ddd_m_units.tot_units_new - ddd_m_units.tot_units_old)/ddd_m_units.tot_units_old)*100
#ddd_m_units
ddd_m_units_deviation = ddd_m_units[((ddd_m_units['deviation'] <= -5) | (ddd_m_units['deviation'] >=5))]
#ddd_m_units_deviation
ddd_m_units_deviations = ddd_m_units_deviation[((ddd_m_units_deviation['tot_units_new'] != 0) & (ddd_m_units_deviation['tot_units_old'] != 0))]
#ddd_m_units_deviations
ddd_m_units_zero_sales = ddd_m_units[(ddd_m_units['tot_units_new'] == 0) | (ddd_m_units['tot_units_old'] == 0)]
#ddd_m_units_zero_sales

ddd_m_units.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\ddd_monthly\ddd_m_units.xlsx",sheet_name = "ddd_m_units")

ddd_m_units_deviations_flag = 'N'
ddd_m_units_zero_sales_flag = 'N'
ddd_m_dollars_deviations_flag = 'N'
ddd_m_dollars_zero_sales_flag = 'N'

if ddd_m_units_deviations.empty == True:
 ddd_m_units_deviations_flag = 'Y'
else:
 ddd_m_units_deviations.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\ddd_monthly\deviations_units.xlsx",sheet_name = "deviations_units")

if ddd_m_units_zero_sales.empty == True:
 ddd_m_units_zero_sales_flag = 'Y'
else:
 ddd_m_units_zero_sales.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\ddd_monthly\zero_sales_units.xlsx",sheet_name = "zero_sales_units")

##Zero_sales_units:

if ddd_m_units_deviations_flag == 'Y':
  pass
else:
 zs_ddd_m_df = pd.read_excel(r"C:\Users\ksxp996\Desktop\unaligned\ddd_monthly\zero_sales_units.xlsx")
 ddd_m_zero_sales_units = zs_ddd_m_df.fillna(method = "ffill")
#ddd_m_zero_sales_units

 ddd_m_zero_sales_units.loc[(ddd_m_zero_sales_units['tot_units_old'] == 0) & (ddd_m_zero_sales_units['tot_units_new'] == 0), 'comments'] = "no comments"
 ddd_m_zero_sales_units.loc[(ddd_m_zero_sales_units['tot_units_new'] == 0) & (ddd_m_zero_sales_units['tot_units_old'] != 0), 'comments'] = ddd_m_zero_sales_units['prod_nm'] + " sales dropped to zero under " + ddd_m_zero_sales_units["mkt_nm"]
 ddd_m_zero_sales_units.loc[(ddd_m_zero_sales_units['tot_units_old'] == 0) & (ddd_m_zero_sales_units['tot_units_new'] != 0), 'comments'] = ddd_m_zero_sales_units['prod_nm'] + " is flowing under " + ddd_m_zero_sales_units["mkt_nm"]

##Deviation_units:

if ddd_m_units_zero_sales_flag == 'Y':
  pass
else:
 dev_ddd_m_df = pd.read_excel(r"C:\Users\ksxp996\Desktop\unaligned\ddd_monthly\deviations_units.xlsx")
 ddd_m_deviation_units = dev_ddd_m_df.fillna(method = "ffill")
#ddd_m_deviation_units

#deviation_units['deviation']= deviation_units['deviation'].astype(str)

 ddd_m_deviation_units.loc[ddd_m_deviation_units['deviation'] > 5, 'comments'] = ddd_m_deviation_units['brd_nm'] + " sales is increased in " + ddd_m_deviation_units["mkt_nm"]
 ddd_m_deviation_units.loc[ddd_m_deviation_units['deviation'] < -5, 'comments'] = ddd_m_deviation_units['brd_nm'] + " sales is dropped in " + ddd_m_deviation_units["mkt_nm"]

#############################################DOLLARS##############################################################################################

ddd_m_dollars = input_ddd_m_df.pivot_table(index = ["mkt_nm","brd_nm","prod_nm"],values = ["tot_dolrs_new","tot_dolrs_old"])
ddd_m_dollars["deviation"] = ((ddd_m_dollars.tot_dolrs_new - ddd_m_dollars.tot_dolrs_old)/ddd_m_dollars.tot_dolrs_old)*100
#ddd_m_dollars
ddd_m_dollars_deviation = ddd_m_dollars[((ddd_m_dollars['deviation'] <= -5) | (ddd_m_dollars['deviation'] >=5))]
#ddd_m_dollars_deviation
ddd_m_dollars_deviations = ddd_m_dollars_deviation[((ddd_m_dollars_deviation['tot_dolrs_new'] != 0) & (ddd_m_dollars_deviation['tot_dolrs_old'] != 0))]
#ddd_m_dollars_deviations
ddd_m_dollars_zero_sales = ddd_m_dollars[(ddd_m_dollars['tot_dolrs_new'] == 0) | (ddd_m_dollars['tot_dolrs_old'] == 0)]
#ddd_m_dollars_zero_sales

ddd_m_dollars.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\ddd_monthly\ddd_m_dollars.xlsx",sheet_name = "ddd_m_dollars")

if ddd_m_dollars_deviations.empty == True:
 ddd_m_dollars_deviations_flag = 'Y'
else:
 ddd_m_dollars_deviations.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\ddd_monthly\deviations_dollars.xlsx",sheet_name = "deviations_dollars")
 
if ddd_m_dollars_zero_sales.empty == True:
 ddd_m_dollars_zero_sales_flag = 'Y'
else:
 ddd_m_dollars_zero_sales.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\ddd_monthly\zero_sales_dollars.xlsx",sheet_name = "zero_sales_dollars")

##Zero_sales_dollars:

if ddd_m_dollars_deviations_flag == 'Y':
  pass
else:
 Zero_sales_dollars_df = pd.read_excel(r"C:\Users\ksxp996\Desktop\unaligned\ddd_monthly\zero_sales_dollars.xlsx")
 ddd_m_zero_sales_dollars = Zero_sales_dollars_df.fillna(method = "ffill")
#ddd_m_zero_sales_dollars

 ddd_m_zero_sales_dollars.loc[(ddd_m_zero_sales_dollars['tot_dolrs_old'] == 0) & (ddd_m_zero_sales_dollars['tot_dolrs_new'] == 0), 'comments'] = "no comments"
 ddd_m_zero_sales_dollars.loc[(ddd_m_zero_sales_dollars['tot_dolrs_new'] == 0) & (ddd_m_zero_sales_dollars['tot_dolrs_old'] != 0), 'comments'] = ddd_m_zero_sales_dollars['prod_nm'] + " sales dropped to zero under " + ddd_m_zero_sales_dollars["mkt_nm"]
 ddd_m_zero_sales_dollars.loc[(ddd_m_zero_sales_dollars['tot_dolrs_old'] == 0) & (ddd_m_zero_sales_dollars['tot_dolrs_new'] != 0), 'comments'] = ddd_m_zero_sales_dollars['prod_nm'] + " is flowing under " + ddd_m_zero_sales_dollars["mkt_nm"]

##Deviation_dollars:

if ddd_m_dollars_zero_sales_flag == 'Y':
  pass
else:
 Deviation_dollars_df = pd.read_excel(r"C:\Users\ksxp996\Desktop\unaligned\ddd_monthly\deviations_dollars.xlsx")
 ddd_m_deviation_dollars = Deviation_dollars_df.fillna(method = "ffill")
#ddd_m_deviation_dollars

#deviation_dollars['deviation']= deviation_dollars['deviation'].astype(str)

 ddd_m_deviation_dollars.loc[ddd_m_deviation_dollars['deviation'] > 5, 'comments'] = ddd_m_deviation_dollars['brd_nm'] + " sales is increased in " + ddd_m_deviation_dollars["mkt_nm"]
 ddd_m_deviation_dollars.loc[ddd_m_deviation_dollars['deviation'] < -5, 'comments'] = ddd_m_deviation_dollars['brd_nm'] + " sales is dropped in " + ddd_m_deviation_dollars["mkt_nm"]
 
#############################Writing output to excel file##################################################################################

if [ddd_m_units_deviations_flag == 'Y'] or [ddd_m_units_zero_sales_flag == 'Y'] or [ddd_m_dollars_deviations_flag == 'Y'] or [ddd_m_dollars_zero_sales_flag == 'Y']:
    with pd.ExcelWriter(r"C:\Users\ksxp996\Desktop\unaligned\ddd_monthly\output\ddd_m_output.xlsx") as writer:    
      ddd_m_units.to_excel(writer,sheet_name = "ddd_m_units")
      ddd_m_dollars.to_excel(writer,sheet_name = "ddd_m_dollars")
    
    if ddd_m_units_deviations_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\ksxp996\Desktop\unaligned\ddd_monthly\output\ddd_m_output.xlsx")
        book.create_sheet("deviation_units")
        sheet = book["deviation_units"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\ksxp996\Desktop\unaligned\ddd_monthly\output\ddd_m_output.xlsx")
        
    else:
        with pd.ExcelWriter(r"C:\Users\ksxp996\Desktop\unaligned\ddd_monthly\output\ddd_m_output.xlsx", engine='openpyxl', mode='a') as writer: 
         ddd_m_deviation_units.to_excel(writer,sheet_name = "deviation_units")

    if ddd_m_units_zero_sales_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\ksxp996\Desktop\unaligned\ddd_monthly\output\ddd_m_output.xlsx")
        book.create_sheet("Zero_sales_units")
        sheet = book["Zero_sales_units"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\ksxp996\Desktop\unaligned\ddd_monthly\output\ddd_m_output.xlsx")
        
    else:
        with pd.ExcelWriter(r"C:\Users\ksxp996\Desktop\unaligned\ddd_monthly\output\ddd_m_output.xlsx", engine='openpyxl', mode='a') as writer: 
         ddd_m_zero_sales_units.to_excel(writer,sheet_name = "Zero_sales_units")
   
    if ddd_m_dollars_deviations_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\ksxp996\Desktop\unaligned\ddd_monthly\output\ddd_m_output.xlsx")
        book.create_sheet("deviation_dollars")
        sheet = book["deviation_dollars"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\ksxp996\Desktop\unaligned\ddd_monthly\output\ddd_m_output.xlsx")
        
    else:
        with pd.ExcelWriter(r"C:\Users\ksxp996\Desktop\unaligned\ddd_monthly\output\ddd_m_output.xlsx", engine='openpyxl', mode='a') as writer: 
         ddd_m_deviation_dollars.to_excel(writer,sheet_name = "deviation_dollars")
  

    if ddd_m_dollars_zero_sales_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\ksxp996\Desktop\unaligned\ddd_monthly\output\ddd_m_output.xlsx")
        book.create_sheet("Zero_sales_dollars")
        sheet = book["Zero_sales_dollars"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\ksxp996\Desktop\unaligned\ddd_monthly\output\ddd_m_output.xlsx")
        
    else:
        with pd.ExcelWriter(r"C:\Users\ksxp996\Desktop\unaligned\ddd_monthly\output\ddd_m_output.xlsx", engine='openpyxl', mode='a') as writer: 
         ddd_m_zero_sales_dollars.to_excel(writer,sheet_name = "Zero_sales_dollars")

ddd_m_folder_path = r'C:\Users\ksxp996\Desktop\unaligned\ddd_monthly\\'

for file_name in listdir(ddd_m_folder_path):
    
    if file_name.endswith('.xlsx'):
        
        os.remove(ddd_m_folder_path + file_name)

##########################################XPO WEEKLY############################################################################################


#########################Nrx##################################################################################################################

xpo_w_nrx = input_xpo_w_df.pivot_table(index = ["mkt_nm","brd_nm","prod_nm"],values = ["nrx_new","nrx_old"])
xpo_w_nrx["deviation"] = ((xpo_w_nrx.nrx_new - xpo_w_nrx.nrx_old)/xpo_w_nrx.nrx_old)*100
#xpo_w_nrx
xpo_w_nrx_deviation = xpo_w_nrx[((xpo_w_nrx['deviation'] <= -5) | (xpo_w_nrx['deviation'] >=5))]
#xpo_w_nrx_deviation
xpo_w_nrx_deviations = xpo_w_nrx_deviation[((xpo_w_nrx_deviation['nrx_new'] != 0) & (xpo_w_nrx_deviation['nrx_old'] != 0))]
#xpo_w_nrx_deviations
xpo_w_nrx_zero_sales = xpo_w_nrx[(xpo_w_nrx['nrx_new'] == 0) | (xpo_w_nrx['nrx_old'] == 0)]
#xpo_w_nrx_zero_sales

xpo_w_nrx.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpo_weekly\xpo_w_nrx.xlsx",sheet_name = "xpo_w_nrx")

xpo_w_nrx_deviations_flag = 'N'
xpo_w_nrx_zero_sales_flag = 'N'
xpo_w_trx_deviations_flag = 'N'
xpo_w_trx_zero_sales_flag = 'N'

if xpo_w_nrx_deviations.empty == True:
 xpo_w_nrx_deviations_flag = 'Y'
else:
 xpo_w_nrx_deviations.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpo_weekly\deviations_nrx.xlsx",sheet_name = "deviations_nrx")

if xpo_w_nrx_zero_sales.empty == True:
 xpo_w_nrx_zero_sales_flag = 'Y'
else:
 xpo_w_nrx_zero_sales.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpo_weekly\zero_sales_nrx.xlsx",sheet_name = "zero_sales_nrx")

##Zero_sales_nrx:

if xpo_w_nrx_zero_sales_flag == 'Y':
  pass
else:
 zs_xpo_w_df = pd.read_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpo_weekly\zero_sales_nrx.xlsx")
 xpo_w_zero_sales_nrx = zs_xpo_w_df.fillna(method = "ffill")
#xpo_w_zero_sales_nrx

 xpo_w_zero_sales_nrx.loc[(xpo_w_zero_sales_nrx['nrx_old'] == 0) & (xpo_w_zero_sales_nrx['nrx_new'] == 0), 'comments'] = "no comments"
 xpo_w_zero_sales_nrx.loc[(xpo_w_zero_sales_nrx['nrx_new'] == 0) & (xpo_w_zero_sales_nrx['nrx_old'] != 0), 'comments'] = xpo_w_zero_sales_nrx['prod_nm'] + " sales dropped to zero under " + xpo_w_zero_sales_nrx["mkt_nm"]
 xpo_w_zero_sales_nrx.loc[(xpo_w_zero_sales_nrx['nrx_old'] == 0) & (xpo_w_zero_sales_nrx['nrx_new'] != 0), 'comments'] = xpo_w_zero_sales_nrx['prod_nm'] + " is flowing under " + xpo_w_zero_sales_nrx["mkt_nm"]

##Deviation_nrx:

if xpo_w_nrx_deviations_flag == 'Y':
  pass
else:
 dev_xpo_w_df = pd.read_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpo_weekly\deviations_nrx.xlsx")
 xpo_w_deviation_nrx = dev_xpo_w_df.fillna(method = "ffill")
#xpo_w_deviation_nrx

#deviation_nrx['deviation']= deviation_nrx['deviation'].astype(str)

 xpo_w_deviation_nrx.loc[xpo_w_deviation_nrx['deviation'] > 5, 'comments'] = xpo_w_deviation_nrx['brd_nm'] + " sales is increased in " + xpo_w_deviation_nrx["mkt_nm"]
 xpo_w_deviation_nrx.loc[xpo_w_deviation_nrx['deviation'] < -5, 'comments'] = xpo_w_deviation_nrx['brd_nm'] + " sales is dropped in " + xpo_w_deviation_nrx["mkt_nm"]

#############################################Trx##############################################################################################

xpo_w_trx = input_xpo_w_df.pivot_table(index = ["mkt_nm","brd_nm","prod_nm"],values = ["trx_new","trx_old"])
xpo_w_trx["deviation"] = ((xpo_w_trx.trx_new - xpo_w_trx.trx_old)/xpo_w_trx.trx_old)*100
#xpo_w_trx
xpo_w_trx_deviation = xpo_w_trx[((xpo_w_trx['deviation'] <= -5) | (xpo_w_trx['deviation'] >=5))]
#xpo_w_trx_deviation
xpo_w_trx_deviations = xpo_w_trx_deviation[((xpo_w_trx_deviation['trx_new'] != 0) & (xpo_w_trx_deviation['trx_old'] != 0))]
#xpo_w_trx_deviations
xpo_w_trx_zero_sales = xpo_w_trx[(xpo_w_trx['trx_new'] == 0) | (xpo_w_trx['trx_old'] == 0)]
#xpo_w_trx_zero_sales

xpo_w_trx.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpo_weekly\xpo_w_trx.xlsx",sheet_name = "xpo_w_trx")

if xpo_w_trx_deviations.empty == True:
 xpo_w_trx_deviations_flag = 'Y'
else:
 xpo_w_trx_deviations.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpo_weekly\deviations_trx.xlsx",sheet_name = "deviations_trx")
 
if xpo_w_trx_zero_sales.empty == True:
 xpo_w_trx_zero_sales_flag = 'Y'
else:
 xpo_w_trx_zero_sales.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpo_weekly\zero_sales_trx.xlsx",sheet_name = "zero_sales_trx")

##Zero_sales_trx:

if xpo_w_trx_zero_sales_flag == 'Y':
  pass
else:
 Zero_sales_trx_df = pd.read_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpo_weekly\zero_sales_trx.xlsx")
 xpo_w_zero_sales_trx = Zero_sales_trx_df.fillna(method = "ffill")
#zero_sales_trx

 xpo_w_zero_sales_trx.loc[(xpo_w_zero_sales_trx['trx_old'] == 0) & (xpo_w_zero_sales_trx['trx_new'] == 0), 'comments'] = "no comments"
 xpo_w_zero_sales_trx.loc[(xpo_w_zero_sales_trx['trx_new'] == 0) & (xpo_w_zero_sales_trx['trx_old'] != 0), 'comments'] = xpo_w_zero_sales_trx['prod_nm'] + " sales dropped to zero under " + xpo_w_zero_sales_trx["mkt_nm"]
 xpo_w_zero_sales_trx.loc[(xpo_w_zero_sales_trx['trx_old'] == 0) & (xpo_w_zero_sales_trx['trx_new'] != 0), 'comments'] = xpo_w_zero_sales_trx['prod_nm'] + " is flowing under " + xpo_w_zero_sales_trx["mkt_nm"]

##Deviation_trx:

if xpo_w_trx_deviations_flag == 'Y':
  pass
else:
 Deviation_trx_df = pd.read_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpo_weekly\deviations_trx.xlsx")
 xpo_w_deviation_trx = Deviation_trx_df.fillna(method = "ffill")
#xpo_w_deviation_trx

#deviation_trx['deviation']= deviation_trx['deviation'].astype(str)

 xpo_w_deviation_trx.loc[xpo_w_deviation_trx['deviation'] > 5, 'comments'] = xpo_w_deviation_trx['brd_nm'] + " sales is increased in " + xpo_w_deviation_trx["mkt_nm"]
 xpo_w_deviation_trx.loc[xpo_w_deviation_trx['deviation'] < -5, 'comments'] = xpo_w_deviation_trx['brd_nm'] + " sales is dropped in " + xpo_w_deviation_trx["mkt_nm"]
 
####################################Wrting the output to excel file################################################################### 	

if [xpo_w_nrx_deviations_flag == 'Y'] or [xpo_w_nrx_zero_sales_flag == 'Y'] or [xpo_w_trx_deviations_flag == 'Y'] or [xpo_w_trx_zero_sales_flag == 'Y']:
    with pd.ExcelWriter(r"C:\Users\ksxp996\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx") as writer:    
      xpo_w_nrx.to_excel(writer,sheet_name = "xpo_w_nrx")
      xpo_w_trx.to_excel(writer,sheet_name = "xpo_w_trx")
    
    if xpo_w_nrx_deviations_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\ksxp996\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx")
        book.create_sheet("deviation_nrx")
        sheet = book["deviation_nrx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\ksxp996\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx")
        
    else:
        with pd.ExcelWriter(r"C:\Users\ksxp996\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx", engine='openpyxl', mode='a') as writer: 
         xpo_w_deviation_nrx.to_excel(writer,sheet_name = "deviation_nrx")

    if xpo_w_nrx_zero_sales_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\ksxp996\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx")
        book.create_sheet("Zero_sales_nrx")
        sheet = book["Zero_sales_nrx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\ksxp996\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx")
        
    else:
        with pd.ExcelWriter(r"C:\Users\ksxp996\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx", engine='openpyxl', mode='a') as writer: 
         xpo_w_zero_sales_nrx.to_excel(writer,sheet_name = "zero_sales_nrx")
   
    if xpo_w_trx_deviations_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\ksxp996\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx")
        book.create_sheet("deviation_trx")
        sheet = book["deviation_trx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\ksxp996\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx")
        
    else:
        with pd.ExcelWriter(r"C:\Users\ksxp996\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx", engine='openpyxl', mode='a') as writer: 
         xpo_w_deviation_trx.to_excel(writer,sheet_name = "deviation_trx")
  

    if xpo_w_trx_zero_sales_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\ksxp996\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx")
        book.create_sheet("Zero_sales_trx")
        sheet = book["Zero_sales_trx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\ksxp996\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx")
        
    else:
        with pd.ExcelWriter(r"C:\Users\ksxp996\Desktop\unaligned\xpo_weekly\output\xpo_w_output.xlsx", engine='openpyxl', mode='a') as writer: 
         xpo_w_zero_sales_trx.to_excel(writer,sheet_name = "Zero_sales_trx")

xpo_w_folder_path = r'C:\Users\ksxp996\Desktop\unaligned\xpo_weekly\\'

for file_name in listdir(xpo_w_folder_path):
    
    if file_name.endswith('.xlsx'):
        
        os.remove(xpo_w_folder_path + file_name)
        
#####################################################XPOPD_WEEKLY################################################################


#########################Nrx##################################################################################################################

xpopd_w_nrx = input_xpopd_w_df.pivot_table(index = ["mkt_nm","brd_nm","prod_nm"],values = ["nrx_new","nrx_old"])
xpopd_w_nrx["deviation"] = ((xpopd_w_nrx.nrx_new - xpopd_w_nrx.nrx_old)/xpopd_w_nrx.nrx_old)*100
#xpopd_w_nrx
xpopd_w_nrx_deviation = xpopd_w_nrx[((xpopd_w_nrx['deviation'] <= -5) | (xpopd_w_nrx['deviation'] >=5))]
#xpopd_w_nrx_deviation
xpopd_w_nrx_deviations = xpopd_w_nrx_deviation[((xpopd_w_nrx_deviation['nrx_new'] != 0) & (xpopd_w_nrx_deviation['nrx_old'] != 0))]
#xpopd_w_nrx_deviations
xpopd_w_nrx_zero_sales = xpopd_w_nrx[(xpopd_w_nrx['nrx_new'] == 0) | (xpopd_w_nrx['nrx_old'] == 0)]
#xpopd_w_nrx_zero_sales

xpopd_w_nrx.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_weekly\xpopd_w_nrx.xlsx",sheet_name = "xpopd_w_nrx")

xpopd_w_nrx_deviations_flag = 'N'
xpopd_w_nrx_zero_sales_flag = 'N'
xpopd_w_trx_deviations_flag = 'N'
xpopd_w_trx_zero_sales_flag = 'N'

if xpopd_w_nrx_deviations.empty == True:
 xpopd_w_nrx_deviations_flag = 'Y'
else:
 xpopd_w_nrx_deviations.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_weekly\deviations_nrx.xlsx",sheet_name = "deviations_nrx")

if xpopd_w_nrx_zero_sales.empty == True:
 xpopd_w_nrx_zero_sales_flag = 'Y'
else:
 xpopd_w_nrx_zero_sales.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_weekly\zero_sales_nrx.xlsx",sheet_name = "zero_sales_nrx")

##Zero_sales_nrx:

if xpopd_w_nrx_zero_sales_flag == 'Y':
  pass
else:
 zs_xpopd_w_df = pd.read_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_weekly\zero_sales_nrx.xlsx")
 xpopd_w_zero_sales_nrx = zs_xpopd_w_df.fillna(method = "ffill")
#xpopd_w_zero_sales_nrx

 xpopd_w_zero_sales_nrx.loc[(xpopd_w_zero_sales_nrx['nrx_old'] == 0) & (xpopd_w_zero_sales_nrx['nrx_new'] == 0), 'comments'] = "no comments"
 xpopd_w_zero_sales_nrx.loc[(xpopd_w_zero_sales_nrx['nrx_new'] == 0) & (xpopd_w_zero_sales_nrx['nrx_old'] != 0), 'comments'] = xpopd_w_zero_sales_nrx['prod_nm'] + " sales dropped to zero under " + xpopd_w_zero_sales_nrx["mkt_nm"]
 xpopd_w_zero_sales_nrx.loc[(xpopd_w_zero_sales_nrx['nrx_old'] == 0) & (xpopd_w_zero_sales_nrx['nrx_new'] != 0), 'comments'] = xpopd_w_zero_sales_nrx['prod_nm'] + " is flowing under " + xpopd_w_zero_sales_nrx["mkt_nm"]

##Deviation_nrx:

if xpopd_w_nrx_deviations_flag == 'Y':
  pass
else:
 dev_xpopd_w_df = pd.read_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_weekly\deviations_nrx.xlsx")
 xpopd_w_deviation_nrx = dev_xpopd_w_df.fillna(method = "ffill")
#xpopd_w_deviation_nrx

#deviation_nrx['deviation']= deviation_nrx['deviation'].astype(str)

 xpopd_w_deviation_nrx.loc[xpopd_w_deviation_nrx['deviation'] > 5, 'comments'] = xpopd_w_deviation_nrx['brd_nm'] + " sales is increased in " + xpopd_w_deviation_nrx["mkt_nm"]
 xpopd_w_deviation_nrx.loc[xpopd_w_deviation_nrx['deviation'] < -5, 'comments'] = xpopd_w_deviation_nrx['brd_nm'] + " sales is dropped in " + xpopd_w_deviation_nrx["mkt_nm"]

#############################################Trx##############################################################################################

xpopd_w_trx = input_xpopd_w_df.pivot_table(index = ["mkt_nm","brd_nm","prod_nm"],values = ["trx_new","trx_old"])
xpopd_w_trx["deviation"] = ((xpopd_w_trx.trx_new - xpopd_w_trx.trx_old)/xpopd_w_trx.trx_old)*100
#xpopd_w_trx
xpopd_w_trx_deviation = xpopd_w_trx[((xpopd_w_trx['deviation'] <= -5) | (xpopd_w_trx['deviation'] >=5))]
#xpopd_w_trx_deviation
xpopd_w_trx_deviations = xpopd_w_trx_deviation[((xpopd_w_trx_deviation['trx_new'] != 0) & (xpopd_w_trx_deviation['trx_old'] != 0))]
#xpopd_w_trx_deviations
xpopd_w_trx_zero_sales = xpopd_w_trx[(xpopd_w_trx['trx_new'] == 0) | (xpopd_w_trx['trx_old'] == 0)]
#xpopd_w_trx_zero_sales

xpopd_w_trx.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_weekly\xpopd_w_trx.xlsx",sheet_name = "xpopd_w_trx")

if xpopd_w_trx_deviations.empty == True:
 xpopd_w_trx_deviations_flag = 'Y'
else:
 xpopd_w_trx_deviations.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_weekly\deviations_trx.xlsx",sheet_name = "deviations_trx")
 
if xpopd_w_trx_zero_sales.empty == True:
 xpopd_w_trx_zero_sales_flag = 'Y'
else:
 xpopd_w_trx_zero_sales.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_weekly\zero_sales_trx.xlsx",sheet_name = "zero_sales_trx")

##Zero_sales_trx:

if xpopd_w_trx_zero_sales_flag == 'Y':
  pass
else:
 Zero_sales_trx_df = pd.read_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_weekly\zero_sales_trx.xlsx")
 xpopd_w_zero_sales_trx = Zero_sales_trx_df.fillna(method = "ffill")
#xpopd_w_zero_sales_trx

 xpopd_w_zero_sales_trx.loc[(xpopd_w_zero_sales_trx['trx_old'] == 0) & (xpopd_w_zero_sales_trx['trx_new'] == 0), 'comments'] = "no comments"
 xpopd_w_zero_sales_trx.loc[(xpopd_w_zero_sales_trx['trx_new'] == 0) & (xpopd_w_zero_sales_trx['trx_old'] != 0), 'comments'] = xpopd_w_zero_sales_trx['prod_nm'] + " sales dropped to zero under " + xpopd_w_zero_sales_trx["mkt_nm"]
 xpopd_w_zero_sales_trx.loc[(xpopd_w_zero_sales_trx['trx_old'] == 0) & (xpopd_w_zero_sales_trx['trx_new'] != 0), 'comments'] = xpopd_w_zero_sales_trx['prod_nm'] + " is flowing under " + xpopd_w_zero_sales_trx["mkt_nm"]

##Deviation_trx:

if xpopd_w_trx_deviations_flag == 'Y':
  pass
else:
 Deviation_trx_df = pd.read_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_weekly\deviations_trx.xlsx")
 xpopd_w_deviation_trx = Deviation_trx_df.fillna(method = "ffill")
#xpopd_w_deviation_trx

#deviation_trx['deviation']= deviation_trx['deviation'].astype(str)

 xpopd_w_deviation_trx.loc[xpopd_w_deviation_trx['deviation'] > 5, 'comments'] = xpopd_w_deviation_trx['brd_nm'] + " sales is increased in " + xpopd_w_deviation_trx["mkt_nm"]
 xpopd_w_deviation_trx.loc[xpopd_w_deviation_trx['deviation'] < -5, 'comments'] = xpopd_w_deviation_trx['brd_nm'] + " sales is dropped in " + xpopd_w_deviation_trx["mkt_nm"]
 
####################################Wrting the output to excel file################################################################### 	

if [xpopd_w_nrx_deviations_flag == 'Y'] or [xpopd_w_nrx_zero_sales_flag == 'Y'] or [xpopd_w_trx_deviations_flag == 'Y'] or [xpopd_w_trx_zero_sales_flag == 'Y']:
    with pd.ExcelWriter(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_weekly\output\xpopd_w_output.xlsx") as writer:    
      xpopd_w_nrx.to_excel(writer,sheet_name = "xpopd_w_nrx")
      xpopd_w_trx.to_excel(writer,sheet_name = "xpopd_w_trx")
    
    if xpopd_w_nrx_deviations_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_weekly\output\xpopd_w_output.xlsx")
        book.create_sheet("deviation_nrx")
        sheet = book["deviation_nrx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_weekly\output\xpopd_w_output.xlsx")
        
    else:
        with pd.ExcelWriter(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_weekly\output\xpopd_w_output.xlsx", engine='openpyxl', mode='a') as writer: 
         xpopd_w_deviation_nrx.to_excel(writer,sheet_name = "deviation_nrx")
        
    if xpopd_w_nrx_zero_sales_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_weekly\output\xpopd_w_output.xlsx")
        book.create_sheet("Zero_sales_nrx")
        sheet = book["Zero_sales_nrx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_weekly\output\xpopd_w_output.xlsx")
        
    else:
        with pd.ExcelWriter(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_weekly\output\xpopd_w_output.xlsx", engine='openpyxl', mode='a') as writer: 
         xpopd_w_zero_sales_nrx.to_excel(writer,sheet_name = "zero_sales_nrx")
    
    if xpopd_w_trx_deviations_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_weekly\output\xpopd_w_output.xlsx")
        book.create_sheet("deviation_trx")
        sheet = book["deviation_trx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_weekly\output\xpopd_w_output.xlsx")
        
    else:
        with pd.ExcelWriter(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_weekly\output\xpopd_w_output.xlsx", engine='openpyxl', mode='a') as writer: 
         xpopd_w_deviation_trx.to_excel(writer,sheet_name = "deviation_trx")
  

    if xpopd_w_trx_zero_sales_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_weekly\output\xpopd_w_output.xlsx")
        book.create_sheet("Zero_sales_trx")
        sheet = book["Zero_sales_trx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_weekly\output\xpopd_w_output.xlsx")
        
    else:
        with pd.ExcelWriter(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_weekly\output\xpopd_w_output.xlsx", engine='openpyxl', mode='a') as writer: 
         xpopd_w_zero_sales_trx.to_excel(writer,sheet_name = "Zero_sales_trx")
        
xpopd_w_folder_path = r'C:\Users\ksxp996\Desktop\unaligned\xpopd_weekly\\'

for file_name in listdir(xpopd_w_folder_path):
    
    if file_name.endswith('.xlsx'):
        
        os.remove(xpopd_w_folder_path + file_name)
        
#######################################################XPOPD_MONTHLY##########################################################


#########################Nrx##################################################################################################################

xpopd_m_nrx = input_xpopd_m_df.pivot_table(index = ["mkt_nm","brd_nm","prod_nm"],values = ["nrx_new","nrx_old"])
xpopd_m_nrx["deviation"] = ((xpopd_m_nrx.nrx_new - xpopd_m_nrx.nrx_old)/xpopd_m_nrx.nrx_old)*100
#xpopd_m_nrx
xpopd_m_nrx_deviation = xpopd_m_nrx[((xpopd_m_nrx['deviation'] <= -5) | (xpopd_m_nrx['deviation'] >=5))]
#xpopd_m_nrx_deviation
xpopd_m_nrx_deviations = xpopd_m_nrx_deviation[((xpopd_m_nrx_deviation['nrx_new'] != 0) & (xpopd_m_nrx_deviation['nrx_old'] != 0))]
#xpopd_m_nrx_deviations
xpopd_m_nrx_zero_sales = xpopd_m_nrx[(xpopd_m_nrx['nrx_new'] == 0) | (xpopd_m_nrx['nrx_old'] == 0)]
#xpopd_m_nrx_zero_sales

xpopd_m_nrx.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_monthly\xpopd_m_nrx.xlsx",sheet_name = "xpopd_m_nrx")

xpopd_m_nrx_deviations_flag = 'N'
xpopd_m_nrx_zero_sales_flag = 'N'
xpopd_m_trx_deviations_flag = 'N'
xpopd_m_trx_zero_sales_flag = 'N'

if xpopd_m_nrx_deviations.empty == True:
 xpopd_m_nrx_deviations_flag = 'Y'
else:
 xpopd_m_nrx_deviations.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_monthly\deviations_nrx.xlsx",sheet_name = "deviations_nrx")

if xpopd_m_nrx_zero_sales.empty == True:
 xpopd_m_nrx_zero_sales_flag = 'Y'
else:
 xpopd_m_nrx_zero_sales.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_monthly\zero_sales_nrx.xlsx",sheet_name = "zero_sales_nrx")

##Zero_sales_nrx:

if xpopd_m_nrx_zero_sales_flag == 'Y':
  pass
else:
 zs_xpopd_m_df = pd.read_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_monthly\zero_sales_nrx.xlsx")
 xpopd_m_zero_sales_nrx = zs_xpopd_m_df.fillna(method = "ffill")
#xpopd_m_zero_sales_nrx

 xpopd_m_zero_sales_nrx.loc[(xpopd_m_zero_sales_nrx['nrx_old'] == 0) & (xpopd_m_zero_sales_nrx['nrx_new'] == 0), 'comments'] = "no comments"
 xpopd_m_zero_sales_nrx.loc[(xpopd_m_zero_sales_nrx['nrx_new'] == 0) & (xpopd_m_zero_sales_nrx['nrx_old'] != 0), 'comments'] = xpopd_m_zero_sales_nrx['prod_nm'] + " sales dropped to zero under " + xpopd_m_zero_sales_nrx["mkt_nm"]
 xpopd_m_zero_sales_nrx.loc[(xpopd_m_zero_sales_nrx['nrx_old'] == 0) & (xpopd_m_zero_sales_nrx['nrx_new'] != 0), 'comments'] = xpopd_m_zero_sales_nrx['prod_nm'] + " is flowing under " + xpopd_m_zero_sales_nrx["mkt_nm"]

##Deviation_nrx:

if xpopd_m_nrx_deviations_flag == 'Y':
  pass
else:
 dev_xpopd_m_df = pd.read_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_monthly\deviations_nrx.xlsx")
 xpopd_m_deviation_nrx = dev_xpopd_m_df.fillna(method = "ffill")
#xpopd_m_deviation_nrx

#xpopd_m_deviation_nrx['deviation']= deviation_nrx['deviation'].astype(str)

 xpopd_m_deviation_nrx.loc[xpopd_m_deviation_nrx['deviation'] > 5, 'comments'] = xpopd_m_deviation_nrx['brd_nm'] + " sales is increased in " + xpopd_m_deviation_nrx["mkt_nm"]
 xpopd_m_deviation_nrx.loc[xpopd_m_deviation_nrx['deviation'] < -5, 'comments'] = xpopd_m_deviation_nrx['brd_nm'] + " sales is dropped in " + xpopd_m_deviation_nrx["mkt_nm"]

#############################################Trx##############################################################################################

xpopd_m_trx = input_xpopd_m_df.pivot_table(index = ["mkt_nm","brd_nm","prod_nm"],values = ["trx_new","trx_old"])
xpopd_m_trx["deviation"] = ((xpopd_m_trx.trx_new - xpopd_m_trx.trx_old)/xpopd_m_trx.trx_old)*100
#xpopd_m_trx
xpopd_m_trx_deviation = xpopd_m_trx[((xpopd_m_trx['deviation'] <= -5) | (xpopd_m_trx['deviation'] >=5))]
#xpopd_m_trx_deviation
xpopd_m_trx_deviations = xpopd_m_trx_deviation[((xpopd_m_trx_deviation['trx_new'] != 0) & (xpopd_m_trx_deviation['trx_old'] != 0))]
#xpopd_m_trx_deviations
xpopd_m_trx_zero_sales = xpopd_m_trx[(xpopd_m_trx['trx_new'] == 0) | (xpopd_m_trx['trx_old'] == 0)]
#xpopd_m_trx_zero_sales

xpopd_m_trx.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_monthly\xpopd_m_trx.xlsx",sheet_name = "xpopd_m_trx")

if xpopd_m_trx_deviations.empty == True:
 xpopd_m_trx_deviations_flag = 'Y'
else:
 xpopd_m_trx_deviations.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_monthly\deviations_trx.xlsx",sheet_name = "deviations_trx")
 
if xpopd_m_trx_zero_sales.empty == True:
 xpopd_m_trx_zero_sales_flag = 'Y'
else:
 xpopd_m_trx_zero_sales.to_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_monthly\zero_sales_trx.xlsx",sheet_name = "zero_sales_trx")

##Zero_sales_trx:

if xpopd_m_trx_zero_sales_flag == 'Y':
  pass
else:
 Zero_sales_trx_df = pd.read_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_monthly\zero_sales_trx.xlsx")
 xpopd_m_zero_sales_trx = Zero_sales_trx_df.fillna(method = "ffill")
#xpopd_m_zero_sales_trx

 xpopd_m_zero_sales_trx.loc[(xpopd_m_zero_sales_trx['trx_old'] == 0) & (xpopd_m_zero_sales_trx['trx_new'] == 0), 'comments'] = "no comments"
 xpopd_m_zero_sales_trx.loc[(xpopd_m_zero_sales_trx['trx_new'] == 0) & (xpopd_m_zero_sales_trx['trx_old'] != 0), 'comments'] = xpopd_m_zero_sales_trx['prod_nm'] + " sales dropped to zero under " + xpopd_m_zero_sales_trx["mkt_nm"]
 xpopd_m_zero_sales_trx.loc[(xpopd_m_zero_sales_trx['trx_old'] == 0) & (xpopd_m_zero_sales_trx['trx_new'] != 0), 'comments'] = xpopd_m_zero_sales_trx['prod_nm'] + " is flowing under " + xpopd_m_zero_sales_trx["mkt_nm"]

##Deviation_trx:

if xpopd_m_trx_deviations_flag == 'Y':
  pass
else:
 Deviation_trx_df = pd.read_excel(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_monthly\deviations_trx.xlsx")
 xpopd_m_deviation_trx = Deviation_trx_df.fillna(method = "ffill")
#xpopd_m_deviation_trx

#deviation_trx['deviation']= deviation_trx['deviation'].astype(str)

 xpopd_m_deviation_trx.loc[xpopd_m_deviation_trx['deviation'] > 5, 'comments'] = xpopd_m_deviation_trx['brd_nm'] + " sales is increased in " + xpopd_m_deviation_trx["mkt_nm"]
 xpopd_m_deviation_trx.loc[xpopd_m_deviation_trx['deviation'] < -5, 'comments'] = xpopd_m_deviation_trx['brd_nm'] + " sales is dropped in " + xpopd_m_deviation_trx["mkt_nm"]
 
####################################Wrting the output to excel file################################################################### 	

if [xpopd_m_nrx_deviations_flag == 'Y'] or [xpopd_m_nrx_zero_sales_flag == 'Y'] or [xpopd_m_trx_deviations_flag == 'Y'] or [xpopd_m_trx_zero_sales_flag == 'Y']:
    with pd.ExcelWriter(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_monthly\output\xpopd_m_output.xlsx") as writer:    
      xpopd_m_nrx.to_excel(writer,sheet_name = "xpopd_m_nrx")
      xpopd_m_trx.to_excel(writer,sheet_name = "xpopd_m_trx")
    
    if xpopd_m_nrx_deviations_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_monthly\output\xpopd_m_output.xlsx")
        book.create_sheet("deviation_nrx")
        sheet = book["deviation_nrx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_monthly\output\xpopd_m_output.xlsx")
        
    else:
        with pd.ExcelWriter(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_monthly\output\xpopd_m_output.xlsx", engine='openpyxl', mode='a') as writer: 
         xpopd_m_deviation_nrx.to_excel(writer,sheet_name = "deviation_nrx")   

    if xpopd_m_nrx_zero_sales_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_monthly\output\xpopd_m_output.xlsx")
        book.create_sheet("Zero_sales_nrx")
        sheet = book["Zero_sales_nrx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_monthly\output\xpopd_m_output.xlsx")
        
    else:
        with pd.ExcelWriter(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_monthly\output\xpopd_m_output.xlsx", engine='openpyxl', mode='a') as writer: 
         xpopd_m_zero_sales_nrx.to_excel(writer,sheet_name = "zero_sales_nrx")
    
    if xpopd_m_trx_deviations_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_monthly\output\xpopd_m_output.xlsx")
        book.create_sheet("deviation_trx")
        sheet = book["deviation_trx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_monthly\output\xpopd_m_output.xlsx")
        
    else:
        with pd.ExcelWriter(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_monthly\output\xpopd_m_output.xlsx", engine='openpyxl', mode='a') as writer: 
         xpopd_m_deviation_trx.to_excel(writer,sheet_name = "deviation_trx")
  

    if xpopd_m_trx_zero_sales_flag == 'Y':
        book = openpyxl.load_workbook(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_monthly\output\xpopd_m_output.xlsx")
        book.create_sheet("zero_sales_trx")
        sheet = book["zero_sales_trx"]
        sheet['A1'] = 'No Major Variance found'
        a1 = sheet['A1']
        a1.font = Font(size=18)
        a1.font = Font(color=colors.BLACK, bold=True)
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        a1.fill = greenFill
        book.save(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_monthly\output\xpopd_m_output.xlsx")
        
    else:
        with pd.ExcelWriter(r"C:\Users\ksxp996\Desktop\unaligned\xpopd_monthly\output\xpopd_m_output.xlsx", engine='openpyxl', mode='a') as writer: 
         xpopd_m_zero_sales_trx.to_excel(writer,sheet_name = "zero_sales_trx")

xpopd_m_folder_path = r'C:\Users\ksxp996\Desktop\unaligned\xpopd_monthly\\'

for file_name in listdir(xpopd_m_folder_path):
    
    if file_name.endswith('.xlsx'):
        
        os.remove(xpopd_m_folder_path + file_name)


# In[ ]:




